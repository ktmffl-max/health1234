#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
운동프로그램 워크북 → 모바일 HTML 변환기

사용법:
    python build_mobile.py                       # 폴더에서 가장 최근 워크북을 자동으로 찾음
    python build_mobile.py 운동프로그램_재편_v14.xlsx
    python build_mobile.py 파일.xlsx -o 어디에저장.html

이 스크립트와 워크북을 같은 폴더에 두고 인자 없이 실행하는 것이 가장 편합니다.
엑셀을 고친 뒤 이 명령만 다시 돌리면 모바일 파일이 갱신됩니다.
종목 추가 · 삭제 · 세트 변경 · 메모 수정 전부 자동으로 따라옵니다.

필요 패키지: openpyxl  (pip install openpyxl)

주의 — 엑셀에서 수정한 뒤에는 반드시 엑셀로 한 번 저장해야 합니다.
      openpyxl은 수식의 '계산된 값'을 읽는데, 그 값은 엑셀이 저장할 때 기록됩니다.
"""

import argparse, datetime, json, re, sys
from pathlib import Path

import voice_build

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl이 필요합니다:  pip install openpyxl")

# 콘솔 코드페이지가 cp949면 '—' 같은 문자에서 출력이 터진다.
# HTML은 이미 쓰인 뒤라 결과물은 멀쩡한데 화면엔 traceback만 남으므로 미리 막는다.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass


# ── 시트 이름 규약 ────────────────────────────────────────────────
# 재택 시트는 'D<일차>_<내용>' (8일 주기) 또는 '<요일>_<내용>' (구 주 단위) 둘 다 잡는다.
# 2026-08-28에 8일 주기로 옮기며 D1~D8 형태가 현행이 됐고, 요일 형태는 옛 워크북 호환용이다.
WEEKDAYS = "월화수목금토일"
DAY_SHEET = re.compile(r"^(?:D(\d+)|([월화수목금토일]))_")
# 복귀판은 주 4회 2분할. 12시간 근무 요일이 유동이라 평일 두 세션은 요일이 아니라
# 슬롯('평A' · '평B')으로만 부른다 — 그 주 근무표를 보고 8시간인 날에 넣는다.
BACK_DAYS = [("평A", "복귀_평일A_가슴"), ("평B", "복귀_평일B_어깨"),
             ("토", "복귀_토_하체"), ("일", "복귀_일_등")]


def home_days(wb):
    """재택 시트를 [(라벨, 시트명)] 로. '복귀_'는 앞이 D숫자도 요일도 아니라 걸리지 않는다.

    8일 주기 시트('D5_당기기B')는 라벨이 '5', 구 요일 시트('목_당기기B')는 '목'이다.
    정렬 키만 갈라두면 나머지 코드는 라벨을 문자열로만 다루므로 손댈 곳이 없다."""
    found = []
    for name in wb.sheetnames:
        m = DAY_SHEET.match(name)
        if m:
            found.append((m.group(1) or m.group(2), name))
    return sorted(found, key=lambda p: int(p[0]) if p[0].isdigit()
                  else WEEKDAYS.index(p[0]))

LIFT_KEYS = {"스쿼트": "squat", "데드리프트": "dead",
             "벤치프레스": "bench", "밀리터리 프레스": "press"}

# 볼륨 램프를 중량보다 몇 주 앞세울지.
# 2026-08-10 결정 — 중량은 2주차 그대로 두고 볼륨만 3주차(138세트)부터 간다.
# 0으로 되돌리면 볼륨과 중량이 같은 주차를 쓴다.
RAMP_WEEK_OFFSET = 1

LIFT_META = {                      # 시작중량_설정 행 / 증량기록 행
    "squat": {"ko": "스쿼트",        "start_row": 7,  "inc_row": 6, "accent": "var(--sq)"},
    "dead":  {"ko": "데드리프트",     "start_row": 8,  "inc_row": 5, "accent": "var(--dl)"},
    "bench": {"ko": "벤치프레스",     "start_row": 9,  "inc_row": 7, "accent": "var(--bp)"},
    "press": {"ko": "밀리터리 프레스", "start_row": 10, "inc_row": 8, "accent": "var(--mp)"},
}


def pick_latest(folder):
    """폴더에서 가장 최근에 수정된 워크북을 고른다. 엑셀 임시파일(~$)은 무시."""
    cands = [p for p in Path(folder).glob("*.xlsx") if not p.name.startswith("~$")]
    if not cands:
        sys.exit(f"{folder} 안에 .xlsx 파일이 없습니다. 워크북을 같은 폴더에 두세요.")
    return max(cands, key=lambda p: p.stat().st_mtime)


def s(v):
    """셀 값을 문자열로. None은 빈 문자열."""
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


# ── 홈 화면 앱(PWA) 자산 ─────────────────────────────────────────
# manifest가 있어야 폰이 이걸 웹페이지가 아니라 앱으로 취급한다.
# 없으면 홈 화면에서 열어도 위에 주소창이 남는다.

BG = (0x15, 0x18, 0x1C)      # --bg   본문 배경과 동일
FG = (0xDD, 0xA5, 0x1B)      # --bp   벤치프레스 노랑. 어두운 배경에서 잘 보인다


def _png(size, px):
    """RGB 픽셀 함수를 8bit PNG 바이트로. 표준 zlib만 사용."""
    import zlib, struct
    raw = bytearray()
    for y in range(size):
        raw.append(0)                       # 필터 타입 None
        for x in range(size):
            raw.extend(px(x, y))

    def chunk(tag, data):
        return (struct.pack(">I", len(data)) + tag + data
                + struct.pack(">I", zlib.crc32(tag + data) & 0xFFFFFFFF))

    return (b"\x89PNG\r\n\x1a\n"
            + chunk(b"IHDR", struct.pack(">IIBBBBB", size, size, 8, 2, 0, 0, 0))
            + chunk(b"IDAT", zlib.compress(bytes(raw), 9))
            + chunk(b"IEND", b""))


def barbell_icon(size):
    """바벨 아이콘. 안드로이드가 원형으로 잘라내도 남도록 중앙 80% 안에 그린다."""
    def rect(x0, y0, x1, y1):
        return (x0 * size, y0 * size, x1 * size, y1 * size)

    shapes = [
        rect(0.16, 0.470, 0.84, 0.530),   # 바
        rect(0.26, 0.330, 0.33, 0.670),   # 안쪽 원판 (좌)
        rect(0.67, 0.330, 0.74, 0.670),   # 안쪽 원판 (우)
        rect(0.20, 0.390, 0.25, 0.610),   # 바깥 원판 (좌)
        rect(0.75, 0.390, 0.80, 0.610),   # 바깥 원판 (우)
    ]

    def px(x, y):
        for x0, y0, x1, y1 in shapes:
            if x0 <= x < x1 and y0 <= y < y1:
                return FG
        return BG

    return _png(size, px)


# Chrome은 manifest만으로는 '홈 화면에 추가'(단순 바로가기)까지만 준다.
# '앱 설치'로 뜨게 하려면 service worker가 있어야 하고, 덤으로 오프라인에서도 열린다.
# 네트워크 우선 — 신호가 있으면 항상 최신을 보고, 끊기면 캐시로 떨어진다.
# 워크북을 고칠 때마다 VERSION이 바뀌어 옛 캐시가 폐기되므로 낡은 표가 남지 않는다.
SW_JS = """\
const VERSION = "__VERSION__";
const CACHE = "workout-" + VERSION;
const ASSETS = ["./", "./index.html", "./manifest.webmanifest",
                "./icon-192.png", "./icon-512.png"];

/* 말소리는 몇 MB짜리 한 덩어리이고, 워크북을 고쳐도 하는 말은 거의 그대로다.
   버전 캐시에 같이 넣으면 표 한 줄 고칠 때마다 폰이 그걸 통째로 다시 받는다.
   그래서 파일 이름에 내용 해시를 박아 따로 보관한다 — 하는 말이 바뀔 때만
   이름이 바뀌고, 그때만 새로 받는다. */
const VOICE_CACHE = "workout-voice";
const VOICE = "__VOICE__";

self.addEventListener("install", e => {
  e.waitUntil(Promise.all([
    caches.open(CACHE).then(c => c.addAll(ASSETS)),
    VOICE ? caches.open(VOICE_CACHE)
              .then(c => c.match(VOICE).then(hit => hit || c.add(VOICE)))
              .catch(() => null)
          : null
  ]).then(() => self.skipWaiting()));
});

self.addEventListener("activate", e => {
  e.waitUntil(caches.keys()
    .then(ks => Promise.all(ks.filter(k => k !== CACHE && k !== VOICE_CACHE)
                              .map(k => caches.delete(k))))
    .then(() => caches.open(VOICE_CACHE))
    .then(c => c.keys().then(rs => Promise.all(      /* 지난 말소리는 버린다 */
      rs.filter(r => !VOICE || !r.url.endsWith(VOICE.slice(2)))
        .map(r => c.delete(r)))))
    .then(() => self.clients.claim()));
});

self.addEventListener("fetch", e => {
  if (e.request.method !== "GET") return;
  /* 말소리만은 캐시 우선이다. 파일 이름이 곧 내용이라 한 번 받으면 다시 받을
     이유가 없고, 네트워크 우선으로 두면 앱을 열 때마다 몇 MB가 샌다. */
  if (e.request.url.indexOf("/voice/") >= 0) {
    e.respondWith(caches.open(VOICE_CACHE).then(c =>
      c.match(e.request).then(hit => hit || fetch(e.request).then(res => {
        c.put(e.request, res.clone());
        return res;
      }))));
    return;
  }
  e.respondWith(
    fetch(e.request).then(res => {
      const copy = res.clone();
      caches.open(CACHE).then(c => c.put(e.request, copy));
      return res;
    }).catch(() => caches.match(e.request).then(r => r || caches.match("./index.html")))
  );
});
"""


def write_pwa_assets(site, html, voice_src):
    """manifest·아이콘·service worker를 배포 폴더에 쓴다. 아이콘은 내용이 고정이라 없을 때만 만든다."""
    manifest = {
        "name": "운동 프로그램",
        "short_name": "운동",
        "start_url": "./",
        "scope": "./",
        "display": "standalone",     # 주소창 없이 전체화면
        "orientation": "portrait",
        "background_color": "#15181C",
        "theme_color": "#15181C",
        "icons": [
            {"src": "icon-192.png", "sizes": "192x192", "type": "image/png",
             "purpose": "any maskable"},
            {"src": "icon-512.png", "sizes": "512x512", "type": "image/png",
             "purpose": "any maskable"},
        ],
    }
    (site / "manifest.webmanifest").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    made = []
    for px_size in (192, 512):
        p = site / f"icon-{px_size}.png"
        if not p.exists():
            p.write_bytes(barbell_icon(px_size))
            made.append(p.name)

    import hashlib
    version = hashlib.sha1(html.encode("utf-8")).hexdigest()[:12]
    (site / "sw.js").write_text(
        SW_JS.replace("__VERSION__", version)
             .replace("__VOICE__", "./" + voice_src if voice_src else ""),
        encoding="utf-8")
    return made, version


def num(v, default=0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def is_continuation(t):
    """소제목은 짧고 마침표가 없다. 길거나 문장으로 끝나면 앞 항목의 뒷줄로 본다."""
    return len(t) > 40 or t.endswith((".", "다", "음", "함"))


def is_ramp_cell(t):
    """메인 리프트 웜업 램프의 중량 칸 — 숫자이거나 '생략'. 안내문이 아니다."""
    return t == "생략" or re.fullmatch(r"\d+(\.\d+)?", t) is not None


def find_header(ws, max_row=40):
    """A열이 '#'인 행 = 종목표 머리글 행."""
    for r in range(1, max_row + 1):
        if s(ws.cell(r, 1).value) == "#":
            return r
    return None


def parse_day(ws):
    """요일 시트 하나를 파싱해 dict로. 재택 · 복귀 양쪽 레이아웃 모두 처리."""
    day = {"t": "", "sub": "", "ex": [], "groups": None, "total": 0,
           "main": None, "notes": [], "pre": []}

    # 제목 — 'A요일 — 가슴' 에서 뒷부분만
    title = s(ws.cell(1, 1).value)
    day["t"] = title.split("—", 1)[1].strip() if "—" in title else title
    day["t"] = re.sub(r"\s*\(.*?\)\s*$", "", day["t"])

    # 메인 리프트 — 'A3: 메인: 벤치프레스'
    for r in (2, 3, 4):
        cell = s(ws.cell(r, 1).value)
        if cell.startswith("메인:"):
            day["main"] = LIFT_KEYS.get(cell.split(":", 1)[1].strip())
            break

    hdr = find_header(ws)
    if hdr is None:
        return day

    # 종목표 위쪽 블록 (웜업 · 관절 준비 등) — 세트 카운트 대상이 아닌 안내문
    pre_head, pre_bul = None, []
    SKIP = ("중량(Kg)", "운동 구성")

    def flush():
        """모아둔 블록을 내보낸다.

        불릿(·)이 한 줄도 없으면 제목만 있는 블록이다. v18의
        '웜업 — 스캡 풀업 …: 팔은 편 채 …' 처럼 한 줄에 다 적은 형태가 여기 해당하는데,
        예전에는 이런 줄이 통째로 버려져 앱에서 웜업이 사라졌다. 콜론에서 갈라
        앞은 제목, 뒤는 내용으로 쓴다."""
        if pre_bul:
            day["pre"].append([pre_head or "웜업", pre_bul])
        elif pre_head:
            head, _, rest = pre_head.partition(":")
            day["pre"].append([head.strip(), [rest.strip()] if rest.strip() else []])

    for rr in range(3, hdr):
        t = s(ws.cell(rr, 1).value)
        if (not t or t in SKIP or t.startswith("메인:") or t.startswith("↑")
                or is_ramp_cell(t)):
            continue
        if t.startswith("·"):
            pre_bul.append(t.lstrip("· ").strip())
        elif pre_bul and is_continuation(t):
            pre_bul[-1] += " " + t
        else:
            flush()
            pre_head, pre_bul = t.strip("[]"), []
    flush()

    # 머리글 → 열 번호 지도 (복귀 시트는 '간접 자극' 열이 없음)
    cols = {}
    for c in range(1, ws.max_column + 1):
        cols[s(ws.cell(hdr, c).value)] = c
    C = lambda name: cols.get(name)

    groups, current, flat = [], None, []
    r = hdr + 1
    while r <= ws.max_row:
        a, b = s(ws.cell(r, 1).value), s(ws.cell(r, C("종목")).value)
        d = s(ws.cell(r, C("중량(Kg)")).value)

        if d == "세트 합계":                      # 표 끝
            day["total"] = int(num(ws.cell(r, C("세트")).value))
            r += 1
            break
        if a and not b:                           # 그룹 머리글 (둔근 / 팔 / 복부)
            current = [a, []]
            groups.append(current)
            r += 1
            continue
        if b:
            item = {
                "n":   int(num(a)) if a else 0,
                "name": b,
                "mg":  s(ws.cell(r, C("근육군")).value),
                "w":   d or "—",
                "s":   int(num(ws.cell(r, C("세트")).value)),
                "r":   s(ws.cell(r, C("횟수")).value),
                # 인터벌 타이머용. 이 열이 없는 워크북(복귀 시트)은 0 = 타이머 없음
                "wk":  int(num(ws.cell(r, C("수행(초)")).value)) if C("수행(초)") else 0,
                "rt":  int(num(ws.cell(r, C("휴식(초)")).value)) if C("휴식(초)") else 0,
                "memo": s(ws.cell(r, C("메모")).value) if C("메모") else "",
                "alt": s(ws.cell(r, C("대체 종목")).value) if C("대체 종목") else "",
                "ind": s(ws.cell(r, C("간접 자극")).value) if C("간접 자극") else "",
                # 이 종목이 몇 주차부터 나오는지. 열이 없는 워크북은 전부 1 = 램프 없음
                "rw": int(num(ws.cell(r, C("복귀 주차")).value, 1)) if C("복귀 주차") else 1,
            }
            # 메인 리프트 중량은 주차에 따라 바뀌므로 상단 카드로 넘긴다
            if day["main"] and item["n"] == 1 and "본세트" in item["name"]:
                item["w"] = "위 참조"
            (current[1] if current else flat).append(item)
        r += 1

    day["groups"] = groups or None
    day["ex"] = flat
    if day["total"] == 0:
        allx = flat + [e for _, g in groups for e in g]
        day["total"] = sum(e["s"] for e in allx)

    # 표 아래 A열 텍스트 = 해설. '·'로 시작하면 항목, 아니면 소제목
    head, bullets = None, []
    for rr in range(r, ws.max_row + 1):
        t = s(ws.cell(rr, 1).value)
        if not t:
            continue
        if t.startswith("·"):
            bullets.append(t.lstrip("· ").strip())
        elif bullets and is_continuation(t):  # 셀에서 줄바꿈된 문장의 뒷부분
            bullets[-1] += " " + t
        else:
            if bullets:                       # 소제목 없이 먼저 나온 항목도 살린다
                day["notes"].append([head or "구성 근거", bullets])
            head, bullets = t.strip("[]"), []
    if bullets:
        day["notes"].append([head or "구성 근거", bullets])

    # 부제 — 2행 안내문이 있으면 사용, 없으면 근육군 요약
    sub = s(ws.cell(2, 1).value)
    mgs, seen = [], set()
    for e in flat + [x for _, g in groups for x in g]:
        if e["mg"] and e["mg"] not in seen:
            seen.add(e["mg"]); mgs.append(e["mg"])
    day["sub"] = f"{sub} · {day['total']}세트" if sub else \
                 f"{' · '.join(mgs[:4])} · {day['total']}세트"
    return day


def parse_check(wb, sheet="세트검산", adj_col=None):
    """근육군별 검산표. adj_col을 주면 그 열을 '현장 일 보정' 등급으로 함께 읽는다.

    재택 시트는 G열이 합산 라벨이라 adj_col을 주지 않고, 복귀 시트만 G열에 보정 등급이 있다."""
    ws = wb[sheet]
    rows = []
    for r in range(5, ws.max_row + 1):
        name = s(ws.cell(r, 1).value)
        if not name or name.startswith("직접 세트 총계") or name == "기준선":
            break
        direct, ind = num(ws.cell(r, 2).value), num(ws.cell(r, 3).value)
        real = num(ws.cell(r, 4).value, direct + ind)
        verdict = s(ws.cell(r, 5).value)
        tag = "ok" if verdict == "충분" else ("lo" if verdict == "하한" else "bad")
        rows.append([name, int(direct), int(ind), int(real), tag, verdict,
                     s(ws.cell(r, 6).value),
                     s(ws.cell(r, adj_col).value) if adj_col else ""])
    return rows


def parse_notes_block(ws, start_row, default_head="메모"):
    out, head, bullets = [], None, []
    for r in range(start_row, ws.max_row + 1):
        t = s(ws.cell(r, 1).value)
        if not t:
            continue
        if t.startswith("·") or t.startswith("→"):
            bullets.append(t.lstrip("·→ ").strip())
        elif bullets and is_continuation(t):
            bullets[-1] += " " + t
        else:
            if bullets:
                out.append([head or default_head, bullets])
            head, bullets = t.strip("[]"), []
    if bullets:
        out.append([head or default_head, bullets])
    return out


def parse_plan(wb):
    """주간계획. 열 위치는 머리글로 찾는다 — 6일판은 '메인 리프트' 열이 끼어들어 세트가 E열로 밀렸다."""
    ws = wb["주간계획"]
    hdr = next((r for r in range(1, 12)
                if s(ws.cell(r, 1).value) in ("일차", "요일")), 4)
    cols = {s(ws.cell(hdr, c).value): c for c in range(1, ws.max_column + 1)}
    c_body = cols.get("내용", 2)
    c_sets = cols.get("세트") or cols.get("예상 세트") or 4

    rows, r = [], hdr + 1
    while r <= ws.max_row:
        d = s(ws.cell(r, 1).value)
        if not d or d == "합계":
            break
        rows.append([d, s(ws.cell(r, c_body).value),
                     int(num(ws.cell(r, c_sets).value))])
        r += 1

    return {"sub": s(ws.cell(2, 1).value), "rows": rows,
            "total": sum(x[2] for x in rows),
            "notes": parse_notes_block(ws, r + 1, "배치 논리")}


def parse_ramp(wb):
    """볼륨램프 — 주차별로 어느 종목이 복귀하는지.

    실제 세트를 정하는 것은 요일 시트 K열('복귀 주차')이고, 이 시트는 그 요약과 근거다.
    시트가 없거나 표를 못 찾으면 None (= 항상 설계 볼륨 100%).
    """
    if "볼륨램프" not in wb.sheetnames:
        return None
    ws = wb["볼륨램프"]
    hdr = next((r for r in range(1, 12) if s(ws.cell(r, 1).value) == "주차"
                and s(ws.cell(r, 2).value) == "총 세트"), None)
    if hdr is None:
        return None

    rows, last = [], hdr
    for r in range(hdr + 1, ws.max_row + 1):
        label = s(ws.cell(r, 1).value)
        if not label:                      # 머리글 아래 빈 줄은 넘기고, 표가 끝나면 멈춘다
            if rows:
                break
            continue
        digits = re.findall(r"\d+", label)
        total = ws.cell(r, 2).value
        if not digits or not isinstance(total, (int, float)):
            break
        # '2주 (현재)' 같은 상태 표기는 떼어낸다 — 어느 주차인지는 앱이 직접 계산해 표시한다
        rows.append({"week": int(digits[0]),
                     "label": re.sub(r"\s*\((?:완료|현재)\)", "", label),
                     "total": int(total), "delta": s(ws.cell(r, 3).value),
                     "items": s(ws.cell(r, 4).value)})
        last = r
    if not rows:
        return None
    return {"sub": s(ws.cell(2, 1).value), "rows": rows,
            "offset": RAMP_WEEK_OFFSET,
            "notes": parse_notes_block(ws, last + 1, "설계 논리")}


def parse_anchor(wb):
    """증량기록 '날짜 (1일차)' 열 — 그 사이클의 1일차 날짜. 달력의 기준점이다.

    여러 줄이 채워져 있으면 가장 늦은 날짜를 쓴다. 중간에 쉬어서 주기가 밀리면
    다음 사이클 줄에 날짜를 적는 것만으로 기준이 다시 맞는다.
    비어 있으면 None — 앱이 '오늘이 몇 일차인지' 직접 묻는다.
    """
    if "증량기록" not in wb.sheetnames:
        return None
    ws = wb["증량기록"]
    hdr = next((r for r in range(1, 20) if s(ws.cell(r, 1).value) == "사이클"), None)
    if hdr is None:
        return None
    c_date = next((c for c in range(1, ws.max_column + 1)
                   if s(ws.cell(hdr, c).value).startswith("날짜")), None)
    if not c_date:
        return None
    best = None
    for r in range(hdr + 1, ws.max_row + 1):
        label = s(ws.cell(r, 1).value)
        if not label:
            if best:
                break
            continue
        m = re.match(r"(\d+)", label)
        v = ws.cell(r, c_date).value
        if m and isinstance(v, datetime.datetime):
            cur = {"date": v.strftime("%Y-%m-%d"), "cycle": int(m.group(1))}
            if best is None or cur["date"] > best["date"]:
                best = cur
    return best


def parse_back_plan(wb):
    ws = wb["복귀_주간계획"]
    hdr = None
    for r in range(1, 12):
        # 주 4회판은 요일이 유동이라 머리글이 '슬롯'이다. 옛 워크북('요일')도 같이 받는다.
        if s(ws.cell(r, 1).value) in ("슬롯", "요일"):
            hdr = r; break
    rows = []
    if hdr:
        for r in range(hdr + 1, hdr + 9):
            d = s(ws.cell(r, 1).value)
            if not d or d == "합계":
                break
            rows.append([d, s(ws.cell(r, 2).value), s(ws.cell(r, 3).value),
                         s(ws.cell(r, 4).value), int(num(ws.cell(r, 5).value))])
    return {"sub": " · ".join(x for x in
                              [s(ws.cell(2, 1).value), s(ws.cell(3, 1).value)] if x),
            "rows": rows, "total": sum(x[4] for x in rows),
            "notes": parse_notes_block(ws, (hdr or 4) + 10, "배치 논리")}


def parse_config(wb):
    st, gr = wb["시작중량_설정"], wb["증량기록"]
    lifts = {}
    for key, m in LIFT_META.items():
        lifts[key] = {
            "ko": m["ko"], "accent": m["accent"],
            "start": num(st.cell(m["start_row"], 4).value),
            "prev":  num(st.cell(m["start_row"], 2).value),
            "inc":   num(gr.cell(m["inc_row"], 3).value),
            "sets":  int(num(re.sub(r"\D+", " ", s(st.cell(m["start_row"], 6).value))
                             .split()[-1], 1)),
        }
    weeks = 0
    for r in range(10, 60):
        if s(gr.cell(r, 1).value).endswith(("사이클", "주차")):
            weeks += 1
        elif weeks:
            break
    return {"lifts": lifts, "weeks": weeks or 12,
            "week": int(num(st.cell(4, 3).value, 1))}


def parse_progress_notes(wb):
    return parse_notes_block(wb["증량기록"], 23, "운용 규칙")


def all_ex(day):
    """요일 하나의 종목 전부. 그룹으로 묶인 시트와 평평한 시트 양쪽을 같이 다룬다."""
    return (day.get("ex") or []) + [x for _, g in (day.get("groups") or []) for x in g]


def build_voice(site, data):
    """'다음 종목' 안내에 쓸 말소리를 굽고 스프라이트 한 덩어리로 묶는다.

    통째로 구울 수 있는 건 통째로 굽는다 — 종목 이름, 근육군, 그리고
    '맨몸'처럼 숫자가 아닌 칸. 반대로 숫자는 조각으로만 굽는다.
    중량은 주차마다 오르고 앱에서도 고칠 수 있어서 미리 구워 둘 수가 없다.
    그래서 0~99와 백 단위, '점오', '킬로 · 세트 · 회 · 에서'를 따로 두고
    앱이 그 순간의 값으로 이어 붙인다 — voice_build.py 머리말 참조.

    SAPI가 없어 못 구우면 지난 번 결과(sprite.json)를 그대로 쓴다.
    """
    say = {}
    for src in ("home", "back"):
        for day in data.get(src, {}).values():
            for e in all_ex(day):
                for raw, fn in ((e.get("name", ""), voice_build.say_name),
                                (e.get("mg", ""), voice_build.say_mg)):
                    if raw and raw not in say:
                        say[raw] = fn(raw)
                # 숫자·범위는 앱이 조립한다. 여기서 굽는 건 예외 표기뿐이다
                for raw, fn in ((str(e.get("r", "")), voice_build.say_reps),
                                (str(e.get("w", "")), voice_build.say_weight)):
                    if (raw and raw not in say and raw != "위 참조"
                            and not voice_build.is_plain_number(raw)):
                        say[raw] = fn(raw)

    for n in range(100):                                  # 0~99
        say.setdefault(voice_build.sino(n), voice_build.sino(n))
    for h in range(1, 10):                                # 백 ~ 구백
        say.setdefault(voice_build.hundreds(h), voice_build.hundreds(h))
    for w in list(voice_build.FRAC.values()) + ["점"] + voice_build.UNITS:
        say.setdefault(w, w)

    say = {k: v for k, v in say.items() if v}
    # 삐 소리 뒤에 붙는 네 마디는 이미 구워져 있다. 다시 굽지 않고 그대로 넣는다 —
    # 새로 구우면 지금 쓰는 말투와 미묘하게 달라진다.
    old = {k: site / "voice" / (k + ".wav") for k in ("prep", "work", "rest", "done")}
    old = {k: v for k, v in old.items() if v.exists()}

    src, offs = voice_build.build(site, say, old)
    keep = site / "voice" / "sprite.json"
    if src:
        voice = {"src": src, "map": offs}
        keep.write_text(json.dumps(voice, ensure_ascii=False), encoding="utf-8")
        return voice
    if keep.exists():          # 이번엔 못 구웠다 — 지난 번 것을 지킨다
        try:
            return json.loads(keep.read_text(encoding="utf-8"))
        except ValueError:
            pass
    return None


def build(xlsx_path, out_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    home = home_days(wb)
    if not home:
        sys.exit("재택 일차 시트를 찾을 수 없습니다. "
                 "시트 이름이 'D1_당기기A' 처럼 'D<일차>_<내용>' 이어야 합니다.")
    missing = [n for _, n in BACK_DAYS if n not in wb.sheetnames]
    if missing:
        sys.exit("시트를 찾을 수 없습니다: " + ", ".join(missing))

    data = {
        "config": parse_config(wb),
        "home":   {label: parse_day(wb[sheet]) for label, sheet in home},
        "back":   {label: parse_day(wb[sheet]) for label, sheet in BACK_DAYS},
        "check":  parse_check(wb),
        "backCheck": parse_check(wb, "복귀_세트검산", adj_col=7),
        "plan":   parse_plan(wb),
        "ramp":   parse_ramp(wb),
        "backPlan": parse_back_plan(wb),
        "progressNotes": parse_progress_notes(wb),
        "anchor": parse_anchor(wb),
        "source": Path(xlsx_path).name,
    }
    # 램프가 실제로 걸려 있는지는 요일 시트 K열이 정한다.
    # 전부 1이면 모든 종목이 1주차부터 나오는 것이므로 앱에 램프를 띄우지 않는다 —
    # 100%만 적힌 표는 매일 보는 화면에서 잡음이다. 시트에는 그대로 남아 있다.
    ret = [e.get("rw", 1) for d in data["home"].values() for e in all_ex(d)]
    if data["ramp"] and ret and max(ret) > 1:
        data["ramp"]["fullWeek"] = max(ret)
    else:
        data["ramp"] = None

    # 휴식일은 주간계획에서 가져와 끼워 넣는다
    for d, content, sets in data["plan"]["rows"]:
        if sets == 0 and d not in data["home"]:
            data["home"][d] = {"t": content or "휴식", "sub": "", "rest": True}
    order = [d for d, _, _ in data["plan"]["rows"]]
    data["homeOrder"] = order or [l for l, _ in home]
    data["backOrder"] = [l for l, _ in BACK_DAYS]

    # 배포용 사본. 파일명을 index.html로 고정해야 사이트 주소 루트가 열리고,
    # 워크북이 v15, v16으로 올라가도 폰 홈화면 바로가기가 안 깨진다.
    site = Path(xlsx_path).resolve().parent / "docs"
    site.mkdir(exist_ok=True)

    # 말소리는 표보다 먼저 만든다 — 어느 구간이 무슨 말인지가 HTML에 실린다
    data["voice"] = build_voice(site, data)

    html = TEMPLATE.replace("/*__DATA__*/ null",
                            json.dumps(data, ensure_ascii=False))
    Path(out_path).write_text(html, encoding="utf-8")
    (site / "index.html").write_text(html, encoding="utf-8")
    made, version = write_pwa_assets(site, html,
                                     data["voice"] and data["voice"]["src"])

    print(f"완료 → {out_path}")
    print(f"       {site / 'index.html'}  (배포용, 캐시버전 {version})")
    if made:
        print(f"       아이콘 생성: {', '.join(made)}")
    train = sum(1 for _, _, n in data["plan"]["rows"] if n)
    print(f"  훈련 {train}일 · 사이클당 {data['plan']['total']}세트 "
          f"(주당 환산 {round(data['plan']['total'] * 7 / 8)}) "
          f"· {data['config']['weeks']}사이클 계획")
    if data["ramp"]:
        vw = data["config"]["week"] + RAMP_WEEK_OFFSET
        print("  볼륨램프(종목 복귀): " + " / ".join(
            f"{r['label']} {r['total']}세트" for r in data["ramp"]["rows"]))
        print(f"       중량 {data['config']['week']}주차 · 볼륨 {vw}주차"
              + (f" (앞세움 +{RAMP_WEEK_OFFSET})" if RAMP_WEEK_OFFSET else ""))
    for row in data["check"]:
        if row[4] != "ok":
            print(f"  [주의] {row[0]}: 실질 {row[3]}세트 — {row[5]}")


# ── HTML 템플릿 ──────────────────────────────────────────────────
TEMPLATE = r"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
<meta name="theme-color" content="#15181C">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<meta name="apple-mobile-web-app-title" content="운동">
<meta name="robots" content="noindex, nofollow">
<link rel="manifest" href="manifest.webmanifest">
<link rel="apple-touch-icon" href="icon-192.png">
<link rel="icon" type="image/png" sizes="512x512" href="icon-512.png">
<title>운동 프로그램</title>
<script>
/* file:// 로 연 로컬 사본에는 navigator.serviceWorker 가 없으므로 조용히 넘어간다 */
if ("serviceWorker" in navigator) {
  addEventListener("load", function () {
    navigator.serviceWorker.register("sw.js").catch(function () {});
  });
}
</script>
<link rel="stylesheet" as="style" crossorigin href="https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/variable/pretendardvariable-dynamic-subset.min.css">
<style>
:root{
  --bg:#15181C; --surface:#1E232A; --raised:#272E37; --line:#333B45;
  --ink:#F0EEE9; --mute:#8B95A1; --dim:#5F6874;
  --sq:#C93A42; --dl:#2B62B0; --bp:#DDA51B; --mp:#2E9161;
  --chrome:#B9C0C8; --pad:16px;
}
*{box-sizing:border-box;-webkit-tap-highlight-color:transparent}
html,body{margin:0;padding:0}
body{background:var(--bg);color:var(--ink);
  font-family:"Pretendard Variable",Pretendard,-apple-system,BlinkMacSystemFont,"Apple SD Gothic Neo","Segoe UI",Roboto,"Noto Sans KR",sans-serif;
  font-size:15px;line-height:1.55;letter-spacing:-.01em;padding-bottom:56px;-webkit-text-size-adjust:100%}
.num{font-variant-numeric:tabular-nums;font-feature-settings:"tnum"}
header{position:sticky;top:0;z-index:40;background:rgba(21,24,28,.94);
  backdrop-filter:saturate(160%) blur(12px);border-bottom:1px solid var(--line);
  padding:calc(env(safe-area-inset-top) + 10px) var(--pad) 0}
.brand{display:flex;align-items:baseline;gap:8px;margin-bottom:10px}
.brand h1{font-size:15px;font-weight:800;margin:0;letter-spacing:-.02em}
.brand span{font-size:11px;color:var(--dim);font-weight:600}
.modes{display:flex;gap:6px;margin-left:auto}
.modes button{font:inherit;font-size:11.5px;font-weight:700;padding:5px 11px;border-radius:999px;
  background:transparent;color:var(--mute);border:1px solid var(--line);cursor:pointer}
.modes button[aria-pressed="true"]{background:var(--ink);color:var(--bg);border-color:var(--ink)}
.weekbar{display:flex;align-items:center;gap:12px;padding-bottom:12px}
.wk-step{width:38px;height:38px;flex:0 0 auto;border-radius:11px;border:1px solid var(--line);
  background:var(--surface);color:var(--ink);font-size:20px;line-height:1;cursor:pointer;
  display:grid;place-items:center;font-family:inherit}
.wk-step:active{background:var(--raised)}
.wk-step[disabled]{opacity:.3}
.wk-read{flex:1;text-align:center}
.wk-read b{font-size:23px;font-weight:800;letter-spacing:-.03em}
.wk-read i{font-style:normal;font-size:11px;color:var(--dim);display:block;margin-top:-2px;letter-spacing:.06em}
.wk-track{display:flex;gap:3px;padding:0 0 10px}
.wk-track span{flex:1;height:3px;border-radius:2px;background:#2A313A}
.wk-track span.on{background:var(--ink)}
nav{display:flex;overflow-x:auto;scrollbar-width:none;border-top:1px solid var(--line);
  margin:0 calc(var(--pad)*-1);padding:0 var(--pad)}
nav::-webkit-scrollbar{display:none}
nav button{font:inherit;font-size:13px;font-weight:700;white-space:nowrap;background:none;
  border:none;border-bottom:2px solid transparent;color:var(--dim);padding:11px 13px;cursor:pointer}
nav button[aria-pressed="true"]{color:var(--ink);border-bottom-color:var(--ink)}
main{padding:18px var(--pad) 0;max-width:640px;margin:0 auto}
.daytitle{font-size:22px;font-weight:800;letter-spacing:-.03em;margin:0 0 3px}
.daysub{font-size:12.5px;color:var(--mute);margin:0 0 18px;line-height:1.5}
.sub-h{font-size:15px;font-weight:800;letter-spacing:-.02em;margin:30px 0 3px}
/* 재적응 구간 알림 — 이 주차에 실제로 수행하는 세트가 설계값과 다르다는 표시 */
.rampbar{display:flex;flex-direction:column;gap:2px;background:var(--surface);
  border:1px solid var(--line);border-left:3px solid var(--bp);border-radius:11px;
  padding:10px 13px;margin:0 0 18px}
.rampbar b{font-size:12.5px;font-weight:800;color:var(--bp)}
.rampbar b span{font-weight:600;color:var(--mute)}
.rampbar span{font-size:11.5px;color:var(--mute);line-height:1.5}
.lift{background:var(--surface);border:1px solid var(--line);border-radius:16px;overflow:hidden;margin-bottom:20px}
.lift-top{padding:15px 16px 13px;display:flex;align-items:flex-end;gap:14px;border-left:4px solid var(--accent)}
.lift-name{flex:1;min-width:0}
.lift-name .eyebrow{font-size:10px;font-weight:800;letter-spacing:.14em;color:var(--accent);margin-bottom:3px}
.lift-name h3{margin:0;font-size:17px;font-weight:800;letter-spacing:-.02em}
.lift-name p{margin:2px 0 0;font-size:12px;color:var(--mute)}
.bigw{text-align:right;line-height:.9}
.bigw b{font-size:42px;font-weight:800;letter-spacing:-.045em}
.bigw i{font-style:normal;font-size:13px;font-weight:700;color:var(--mute);margin-left:2px}
.bar{padding:0 16px 14px;border-left:4px solid var(--accent)}
.bar-label{font-size:10px;font-weight:800;letter-spacing:.12em;color:var(--dim);margin-bottom:7px}
.stack{display:flex;align-items:center;gap:3px;min-height:44px}
.sleeve{width:26px;height:5px;background:var(--chrome);border-radius:2px;opacity:.55}
.pl{border-radius:2px;flex:0 0 auto}
.stack em{font-style:normal;font-size:12px;color:var(--dim);margin-left:6px}
.ramp{border-top:1px solid var(--line);display:grid;grid-template-columns:1fr auto auto;font-size:13.5px}
.ramp div{padding:9px 16px;border-bottom:1px solid rgba(51,59,69,.5)}
.ramp .h{font-size:10px;font-weight:800;letter-spacing:.1em;color:var(--dim);padding-top:10px;padding-bottom:6px}
.ramp .r-w{font-weight:700}
.ramp .r-x{color:var(--mute);text-align:right;font-size:12.5px}
.ramp .skip{color:var(--dim);font-weight:500}
.ramp .work{background:var(--raised)}
.ramp .work .r-w{color:var(--accent)}
.ramp>div:nth-last-child(-n+3){border-bottom:none}
.grouphead{font-size:10.5px;font-weight:800;letter-spacing:.14em;color:var(--dim);margin:22px 0 9px}
.ex{background:var(--surface);border:1px solid var(--line);border-radius:13px;padding:13px 14px;margin-bottom:9px}
.ex-h{display:flex;gap:10px;align-items:flex-start}
.ex-n{font-size:11px;font-weight:800;color:var(--dim);width:15px;flex:0 0 auto;padding-top:2px}
.ex-t{flex:1;min-width:0}
.ex-t h4{margin:0;font-size:15.5px;font-weight:700;letter-spacing:-.02em}
.ex-t .mg{font-size:11.5px;color:var(--mute);margin-top:1px}
.ex-p{text-align:right;flex:0 0 auto}
.ex-p b{font-size:16px;font-weight:800;letter-spacing:-.02em;display:block}
.ex-p span{font-size:11.5px;color:var(--mute)}
/* 이번 주에 쉬는 종목 — 흐리게 두고 언제 돌아오는지만 남긴다 */
.ex.hold{opacity:.5}
/* 설계 세트 0(프로그램에서 빼둔 종목) — 흐름에서 빼고 하단 접기 안에 모은다 */
.holdbox{margin-top:12px;padding:11px 14px;border:1px dashed var(--line);border-radius:13px}
.holdbox>summary{color:var(--mute)}
.holdbox>summary i{font-style:normal;font-weight:600;color:var(--dim)}
.holdbox[open]>summary{margin-bottom:11px}
.holdbox .ex{margin-bottom:0}
.holdbox .ex+.ex{margin-top:8px}
.ex-p b.hold-w{font-size:12px;font-weight:800;letter-spacing:0;color:var(--bp)}
.ex-memo{font-size:12.5px;color:var(--mute);margin-top:9px;padding-top:9px;border-top:1px solid var(--line);line-height:1.5}
details{margin-top:9px}
details summary{list-style:none;font-size:11.5px;font-weight:700;color:var(--dim);cursor:pointer;display:flex;align-items:center;gap:5px}
details summary::-webkit-details-marker{display:none}
details summary::after{content:"＋";font-size:12px}
details[open] summary::after{content:"−"}
details p{margin:7px 0 0;font-size:12.5px;color:var(--mute);line-height:1.6}
.total{display:flex;justify-content:space-between;align-items:baseline;padding:12px 14px;
  border:1px dashed var(--line);border-radius:13px;margin-top:12px}
.total span{font-size:11px;font-weight:800;letter-spacing:.12em;color:var(--dim)}
.total b{font-size:19px;font-weight:800}
.notes{margin:22px 0 0}
.notes h5{font-size:12.5px;font-weight:800;margin:0 0 8px}
.notes ul{margin:0 0 18px;padding:0;list-style:none}
.notes li{font-size:12.5px;color:var(--mute);line-height:1.65;padding-left:12px;position:relative;margin-bottom:5px}
.notes li::before{content:"";position:absolute;left:0;top:9px;width:4px;height:1px;background:var(--dim)}
.tbl{width:100%;border-collapse:collapse;font-size:12.5px;margin-bottom:6px}
.tbl th{font-size:10px;font-weight:800;letter-spacing:.08em;color:var(--dim);text-align:right;
  padding:0 0 8px;border-bottom:1px solid var(--line);white-space:nowrap}
.tbl th:first-child{text-align:left}
.tbl td{padding:9px 0;border-bottom:1px solid rgba(51,59,69,.45);text-align:right;white-space:nowrap}
.tbl td:first-child{text-align:left;white-space:normal;font-weight:600;padding-right:8px}
.tbl td.wrap{white-space:normal;text-align:left;color:var(--mute);font-size:11.5px;padding-left:8px}
/* 오른쪽 정렬 열 바로 뒤에 오는 왼쪽 정렬 머리글은 여백이 없으면 앞 글자와 붙는다 */
.tbl th.wrap{text-align:left;padding-left:8px}
.judge{font-size:10.5px;font-weight:800;padding:2px 7px;border-radius:999px;border:1px solid;
  white-space:nowrap;display:inline-block}
.j-ok{color:var(--mp);border-color:rgba(46,145,97,.45)}
.j-lo{color:var(--bp);border-color:rgba(221,165,27,.45)}
.j-bad{color:var(--sq);border-color:rgba(201,58,66,.45)}
.footnote{font-size:11.5px;color:var(--dim);margin-top:18px;line-height:1.6}
.rest{opacity:.42}
[data-lift]{cursor:pointer}
.ex-w[data-exkey]{border-bottom:1px dotted var(--dim);cursor:pointer;padding-bottom:1px}
.ex-w.edited{color:var(--bp);border-bottom-color:var(--bp)}
.bigw.edited b{color:var(--bp)}
.ovnote{display:block;font-size:10px;font-weight:700;color:var(--bp);margin-top:5px;line-height:1.3}
.w-input{font:inherit;font-weight:700;background:var(--raised);color:var(--ink);
  border:1px solid var(--dim);border-radius:8px;padding:5px 8px;width:96px;
  text-align:right;font-size:16px}
.bigw .w-input{width:120px;font-size:22px}
.ovbox{margin-top:22px;border:1px dashed var(--bp);border-radius:13px;padding:13px 14px}
.ovbox h5{font-size:12.5px;font-weight:800;margin:0;color:var(--bp)}
.ovbox h5 span{font-weight:600;color:var(--mute);font-size:11px}
.ovbox ul{margin:8px 0 11px;padding:0;list-style:none}
.ovbox li{font-size:12.5px;color:var(--mute);line-height:1.65}
.ovbox button{font:inherit;font-size:12px;font-weight:700;padding:7px 12px;border-radius:9px;
  background:transparent;color:var(--sq);border:1px solid var(--sq);cursor:pointer}

/* ── 달력 ──────────────────────────────────────────────────────── */
.cal-bar{display:flex;align-items:center;gap:8px;margin:2px 0 12px}
.cal-bar h3{flex:1;margin:0;font-size:17px;font-weight:800;letter-spacing:-.02em}
.cal-bar button{font:inherit;font-size:13px;font-weight:700;padding:7px 12px;border-radius:9px;
  background:var(--surface);color:var(--ink);border:1px solid var(--line);cursor:pointer}
.cal-grid{display:grid;grid-template-columns:repeat(7,1fr);gap:3px}
.cal-dow{text-align:center;font-size:11px;font-weight:800;color:var(--dim);padding:2px 0 4px}
.cal-dow.sun{color:var(--sq)}
.cal-cell{min-height:62px;border-radius:8px;padding:5px 4px;background:var(--surface);
  border:1px solid transparent;border-left:3px solid var(--acc,var(--line));
  display:flex;flex-direction:column;gap:1px;cursor:pointer;overflow:hidden}
.cal-cell.pad,.cal-cell.blank{background:transparent;border-color:transparent;cursor:default}
.cal-cell.blank b{color:var(--dim);font-weight:600}
.cal-cell.rest{background:transparent;border-left-color:var(--line);border:1px dashed var(--line)}
.cal-cell.today{border-color:var(--ink);background:var(--raised)}
.cal-cell b{font-size:13px;font-weight:800;font-variant-numeric:tabular-nums;line-height:1.1}
.cal-cell.sun b{color:var(--sq)}
.cal-cell i{font-style:normal;font-size:9.5px;font-weight:800;color:var(--acc,var(--mute));
  letter-spacing:-.02em}
.cal-cell span{font-size:9.5px;line-height:1.25;color:var(--mute);word-break:keep-all}
.cal-cell em{font-style:normal;font-size:9px;font-weight:800;color:var(--dim);margin-top:auto}
.cal-key{display:flex;flex-wrap:wrap;gap:5px 12px;margin:13px 2px 0}
.cal-key span{display:inline-flex;align-items:center;gap:5px;font-size:11px;
  font-weight:700;color:var(--mute)}
.cal-key i{width:9px;height:9px;border-radius:3px;flex:none}
.cal-key b{font-weight:700;color:var(--dim)}
.cal-now{background:var(--surface);border:1px solid var(--line);border-radius:12px;
  padding:13px 14px;margin:0 0 13px}
.cal-now b{display:block;font-size:19px;font-weight:800;letter-spacing:-.02em}
.cal-now p{margin:3px 0 0;font-size:12.5px;color:var(--mute)}
.cal-fix{margin-top:11px}
.cal-fix summary{font-size:12px;font-weight:700;color:var(--mute);cursor:pointer}
.cal-days{display:flex;flex-wrap:wrap;gap:6px;margin-top:9px}
.cal-days button{font:inherit;font-size:12px;font-weight:700;padding:8px 11px;border-radius:9px;
  background:var(--raised);color:var(--ink);border:1px solid var(--line);cursor:pointer}
.cal-days button.on{background:var(--ink);color:var(--bg);border-color:var(--ink)}

/* ── 인터벌 타이머 ─────────────────────────────────────────────── */
.tm-go{display:inline-flex;align-items:center;gap:7px;font:inherit;font-size:12px;font-weight:700;
  padding:6px 12px;border-radius:999px;background:var(--raised);color:var(--ink);
  border:1px solid var(--line);cursor:pointer}
.tm-go b{font-variant-numeric:tabular-nums;color:var(--mute);font-weight:600}
.ex-go{margin:10px 0 0}
.tm-none{display:inline-block;font-size:11.5px;font-weight:600;color:var(--dim);
  padding:5px 10px;border-radius:999px;border:1px dashed var(--line)}
.day-go i{font-style:normal;font-weight:600;opacity:.8}
.day-go{display:block;width:100%;margin:0 0 15px;padding:14px;border:0;border-radius:13px;
  background:var(--mp);color:#fff;font:inherit;font-size:15px;font-weight:800;cursor:pointer}
#tm{position:fixed;inset:0;z-index:200;background:var(--bg);display:flex;flex-direction:column;
  padding:calc(env(safe-area-inset-top) + 12px) 18px calc(env(safe-area-inset-bottom) + 18px)}
#tm[hidden]{display:none}
.tm-top{display:flex;align-items:center;justify-content:space-between;gap:10px;
  font-size:12.5px;font-weight:700;color:var(--mute)}
.tm-top button{font:inherit;font-size:19px;line-height:1;background:none;border:0;
  color:var(--mute);padding:2px 6px;cursor:pointer}
.tm-btns{display:flex;align-items:center;gap:2px;flex:none}
#tm-mute.off{opacity:.4}
.tm-body{flex:1;display:flex;flex-direction:column;align-items:center;justify-content:center;
  text-align:center;min-height:0}
#tm-name{margin:0;font-size:25px;font-weight:800;line-height:1.25;letter-spacing:-.02em}
#tm-meta{margin:5px 0 18px;font-size:14px;color:var(--mute);font-variant-numeric:tabular-nums}
.tm-dial{width:min(72vw,286px);aspect-ratio:1;border-radius:50%;display:flex;flex-direction:column;
  align-items:center;justify-content:center;background:var(--surface);
  border:7px solid var(--phase,var(--line));transition:border-color .25s}
.tm-dial span{font-size:15px;font-weight:800;letter-spacing:.03em;color:var(--phase,var(--mute))}
.tm-dial b{font-size:clamp(46px,15vw,66px);font-weight:800;letter-spacing:-.03em;
  font-variant-numeric:tabular-nums;line-height:1.15}
#tm-set{margin:18px 0 0;font-size:15px;font-weight:700;color:var(--mute)}
.tm-ctl{display:grid;grid-template-columns:1fr 1.25fr 1fr;gap:8px;margin-top:10px}
.tm-ctl button{font:inherit;font-size:14px;font-weight:700;padding:16px 4px;border-radius:12px;
  background:var(--surface);color:var(--ink);border:1px solid var(--line);cursor:pointer}
.tm-ctl button.primary{background:var(--ink);color:var(--bg);border-color:var(--ink)}
.tm-up{margin:13px 0 0;text-align:center;font-size:12.5px;color:var(--dim);
  line-height:1.55;min-height:3.2em}   /* 두 줄분을 늘 비워 둔다 — 종목이 바뀔 때
                                          한 줄이 늘어나며 다이얼이 들썩이지 않게 */
.tm-up b{color:var(--mute);font-weight:700;font-size:13.5px}
.tm-up i{display:block;font-style:normal;font-variant-numeric:tabular-nums}
</style>
</head>
<body>
<header>
  <div class="brand">
    <h1>운동 프로그램</h1><span class="num" id="ver"></span>
    <div class="modes">
      <button id="m-home" aria-pressed="true">재택</button>
      <button id="m-back" aria-pressed="false">복귀</button>
    </div>
  </div>
  <div class="weekbar" id="weekbar">
    <button class="wk-step" id="wk-down" aria-label="이전 사이클">−</button>
    <div class="wk-read"><b class="num" id="wk-num">1</b><i id="wk-of">CYCLE</i></div>
    <button class="wk-step" id="wk-up" aria-label="다음 사이클">+</button>
  </div>
  <div class="wk-track" id="wk-track"></div>
  <nav id="nav"></nav>
</header>
<main id="app"></main>
<div id="tm" hidden>
  <div class="tm-top"><span id="tm-ctx"></span><span class="tm-btns"
    ><button id="tm-mute" aria-label="음성 켜기/끄기">&#128266;</button
    ><button id="tm-x" aria-label="타이머 닫기">&#10005;</button></span></div>
  <div class="tm-body">
    <h3 id="tm-name"></h3>
    <p id="tm-meta"></p>
    <div class="tm-dial" id="tm-dial"><span id="tm-phase"></span><b id="tm-time">00:00</b></div>
    <p id="tm-set"></p>
  </div>
  <div class="tm-ctl">
    <button id="tm-prev">&#9664; 이전</button>
    <button id="tm-play" class="primary">일시정지</button>
    <button id="tm-next">건너뛰기 &#9654;</button>
  </div>
  <p class="tm-up" id="tm-up"></p>
</div>
<script>
const D = /*__DATA__*/ null;

const R = v => Math.round(v/2.5)*2.5;
const fmt = v => (Math.round(v*1000)/1000).toString();
const WEEKS = D.config.weeks;
const LIFT = D.config.lifts;

/* ── 이 기기에서만 적용되는 중량 덮어쓰기 ─────────────────────────
   엑셀 원본은 건드리지 않는다. {수정값, 그때의 엑셀값}을 localStorage에 두고,
   엑셀 쪽 값이 바뀐 채 재배포되면 해당 덮어쓰기는 자동으로 버린다. */
const OVK = "wk-ov";
let OV = {lifts:{}, ex:{}};
try{ const o=JSON.parse(localStorage.getItem(OVK)||"null"); if(o&&o.lifts&&o.ex) OV=o; }catch(e){}
function exByKey(key){
  const p=key.split("|"), day=(p[0]==="home"?D.home:D.back)[p[1]];
  if(!day||day.rest) return null;
  const all=(day.ex||[]).concat(day.groups?day.groups.reduce((a,g)=>a.concat(g[1]),[]):[]);
  return all.find(x=>x.name===p[2])||null;
}
const saveOV = () => { try{ localStorage.setItem(OVK, JSON.stringify(OV)); }catch(e){} };
let ovStale=false;
for(const k in OV.lifts){ if(!LIFT[k]||LIFT[k].start!==OV.lifts[k].base){ delete OV.lifts[k]; ovStale=true; } }
for(const k in OV.ex){ const e=exByKey(k); if(!e||e.w!==OV.ex[k].base){ delete OV.ex[k]; ovStale=true; } }
if(ovStale) saveOV();
const startOf = k => OV.lifts[k] ? OV.lifts[k].start : LIFT[k].start;
const workWeight = (k,w) => startOf(k) + LIFT[k].inc*(w-1);

/* ── 볼륨 램프 — 종목 복귀 방식 ───────────────────────────────────
   계수로 전 종목을 한꺼번에 깎던 방식을 버렸다. 종목마다 '복귀 주차'(엑셀 K열)가 있고,
   그 주차 전에는 종목 자체가 빠진다. 나오는 종목은 설계 세트를 그대로 한다.
   엑셀 J열과 같은 식: IF(볼륨 주차 >= 복귀 주차, 설계 세트, 0).
   볼륨 주차는 중량 주차보다 offset만큼 앞서간다 (2026-08-10: 중량 2주차 · 볼륨 3주차).
   복귀 모드에는 주차 개념이 없으므로 적용하지 않는다. */
const RAMP = D.ramp;
const rampOn = () => !!(RAMP && mode==="home");
const vWeek = w => w + (RAMP ? RAMP.offset : 0);
const setsOf = (e,w) => vWeek(w) >= (e.rw||1) ? e.s : 0;
const rampRow = w => RAMP.rows.filter(r=>r.week<=vWeek(w)).pop() || RAMP.rows[0];

function ramp(work){
  const pct=[.4,.6,.75,.9], reps=[5,5,3,2], sets=[2,1,1,1];
  const raw = pct.map(p=>Math.max(20,R(work*p)));
  return raw.map((w,i)=>({w:(i>0 && w<=raw[i-1])?null:w, reps:reps[i], sets:sets[i]}));
}
const PLATES=[{kg:25,c:"#C93A42",h:42,w:11},{kg:20,c:"#2B62B0",h:42,w:11},
  {kg:15,c:"#DDA51B",h:40,w:9},{kg:10,c:"#2E9161",h:37,w:8},
  {kg:5,c:"#E8E6E1",h:30,w:8},{kg:2.5,c:"#3D454F",h:24,w:7},
  {kg:1.25,c:"#B9C0C8",h:19,w:6},{kg:1.125,c:"#8B95A1",h:16,w:6}];
function load(total){
  let side=(total-20)/2, out=[];
  if(side<0.001) return out;
  for(const p of PLATES){ while(side>=p.kg-0.0005){ out.push(p); side=Math.round((side-p.kg)*1000)/1000; } }
  return out;
}
function stackHTML(total){
  const ps=load(total);
  const body = ps.length
    ? ps.map(p=>`<div class="pl" style="background:${p.c};width:${p.w}px;height:${p.h}px"></div>`).join("")
      + `<em>한쪽 ${fmt((total-20)/2)}kg</em>`
    : `<em>빈 봉 20kg</em>`;
  return `<div class="stack"><div class="sleeve"></div>${body}</div>`;
}
function liftCard(key,week,sets){
  const L=LIFT[key]; if(!L) return "";
  const work=workWeight(key,week), rows=ramp(work);
  /* 본세트 수는 종목표 1번 행이 정한다 — 메인 리프트는 어느 주차에도 빠지지 않는다 */
  const ws=(sets==null?L.sets:sets);
  const rp=rows.map(r=>`
    <div class="r-w ${r.w?'':'skip'}">${r.w?fmt(r.w)+'<span style="font-size:11px;color:var(--mute)"> kg</span>':'생략'}</div>
    <div class="r-x num">${r.reps}회</div><div class="r-x num">${r.sets}세트</div>`).join("");
  return `<section class="lift" style="--accent:${L.accent}">
    <div class="lift-top">
      <div class="lift-name"><div class="eyebrow">MAIN LIFT</div><h3>${L.ko}</h3>
        <p class="num">5년 전 ${fmt(L.prev)}kg · 본세트 5회 × ${ws}세트${ws!==L.sets?` <span style="color:var(--dim)">(설계 ${L.sets})</span>`:""}</p></div>
      <div class="bigw ${OV.lifts[key]?'edited':''}" data-lift="${key}"><b class="num">${fmt(work)}</b><i>kg</i>${
        OV.lifts[key]?`<span class="ovnote">수정됨 · 엑셀 기준 ${fmt(LIFT[key].start+LIFT[key].inc*(week-1))}kg</span>`:""}</div>
    </div>
    <div class="bar"><div class="bar-label">한쪽 원판 구성</div>${stackHTML(work)}</div>
    <div class="ramp">
      <div class="h">웜업 램프</div><div class="h" style="text-align:right">반복</div><div class="h" style="text-align:right">세트</div>
      ${rp}
      <div class="work r-w">${fmt(work)}<span style="font-size:11px;color:var(--mute)"> kg</span></div>
      <div class="work r-x num">5회</div><div class="work r-x num">${ws}세트</div>
    </div></section>`;
}
const esc = t => String(t).replace(/&/g,"&amp;").replace(/</g,"&lt;");
const escAttr = t => esc(t).replace(/"/g,"&quot;");
function exHTML(e,prefix,sets,timed){
  const key=prefix+"|"+e.name, ov=OV.ex[key];
  const w=ov?ov.w:e.w, canEdit=e.w!=="위 참조";
  const n=(sets==null?e.s:sets);
  /* 이번 주에 안 하는 종목 — 목록에서 빼지 않고 흐리게 남긴다.
     사라진 게 아니라 몇 주차에 돌아오는지가 보여야 한다. 메모·대체 종목은 그때 읽으면 된다.
     설계 세트(e.s)까지 0이면 램프 대기가 아니라 프로그램에서 빼둔 종목이다 —
     엑셀도 조건부서식($E=0)으로 같은 줄을 연회색 처리한다. 되살릴 수 있게 행은 남긴다. */
  if(n===0) return `<article class="ex hold"><div class="ex-h">
      <div class="ex-n num">${e.n}</div>
      <div class="ex-t"><h4>${esc(e.name)}</h4><div class="mg">${esc(e.mg)}</div></div>
      <div class="ex-p">${e.s===0 ? `<b class="hold-w">보류</b><span class="num">${esc(e.r)}</span>`
        : `<b class="hold-w">${e.rw}주차 복귀</b><span class="num">설계 ${e.s} × ${esc(e.r)}</span>`}
      </div></div></article>`;
  return `<article class="ex"><div class="ex-h">
      <div class="ex-n num">${e.n}</div>
      <div class="ex-t"><h4>${esc(e.name)}</h4><div class="mg">${esc(e.mg)}${e.ind?` · 간접 ${esc(e.ind)}`:""}</div></div>
      <div class="ex-p"><b class="num">${n} × ${esc(e.r)}</b><span class="num ex-w ${ov?'edited':''}"${canEdit?` data-exkey="${escAttr(key)}"`:""}>${esc(w)}</span></div></div>
    ${e.wk?`<div class="ex-go"><button class="tm-go" data-tm="${escAttr(e.name)}">&#9654; 타이머<b>${e.wk}초 · 휴식 ${mmss(e.rt)}</b></button></div>`
      :timed?`<div class="ex-go"><span class="tm-none">타이머 없음 — 엑셀 G·H열(수행 초 · 휴식 초)이 비어 있다</span></div>`:""}
    ${e.memo?`<div class="ex-memo">${esc(e.memo)}</div>`:""}
    ${e.alt?`<details><summary>대체 종목</summary><p>${esc(e.alt)}</p></details>`:""}</article>`;
}
const notesHTML = ns => !ns||!ns.length ? "" :
  `<div class="notes">`+ns.map(([h,ls])=>`<h5>${esc(h)}</h5><ul>${ls.map(l=>`<li>${esc(l)}</li>`).join("")}</ul>`).join("")+`</div>`;

const allEx = d => (d.ex||[]).concat(d.groups?d.groups.reduce((a,g)=>a.concat(g[1]),[]):[]);

function rampBar(all,S,total,done,week){
  const wait={};
  all.forEach(e=>{ if(e.s>0 && S(e)===0) wait[e.rw]=(wait[e.rw]||0)+1; });
  const when=Object.keys(wait).map(Number).sort((a,b)=>a-b)
    .map(k=>`${k}주차에 ${wait[k]}종목`).join(" · ");
  return `<div class="rampbar"><b>볼륨 ${vWeek(week)}주차 · ${done}세트 <span>/ 설계 ${total}</span></b>
    <span>${when?`대기 중 — ${when} 복귀.`:""} 나오는 종목은 설계 세트 그대로 한다</span></div>`;
}

function dayHTML(d,week){
  if(d.rest) return `<h2 class="daytitle">${esc(d.t)}</h2><p class="daysub">${esc(d.sub||"휴식일")}</p>`;
  const prefix=mode+"|"+tab;
  const on=rampOn();
  const S = e => on ? setsOf(e,week) : e.s;
  const all=allEx(d), done=all.reduce((a,e)=>a+S(e),0);
  /* 타이머가 실제로 걸 수 있는 종목 — 수행 초가 있어야 한다 */
  const timed = allEx(d).filter(e=>e.s!==0&&S(e)>0&&e.wk>0);
  const one = e => exHTML(e,prefix,S(e),timed.length>0);
  /* 설계 세트 0 = 프로그램에서 빼둔 종목(옮겼거나 제외했거나).
     목록 흐름에서 빼고 하단 접기에 모은다 — 폰에서는 오늘 할 것만 보이고,
     엑셀 행은 그대로라 세트수만 되돌리면 다시 흐름으로 돌아온다.
     램프 대기(설계는 있는데 이번 주 0)는 몇 주차 복귀인지가 정보라 그대로 남긴다. */
  const held = all.filter(e=>e.s===0), live = list => list.filter(e=>e.s!==0);
  const body = d.groups
    ? d.groups.map(([g,list])=>live(list).length
        ? `<div class="grouphead">${esc(g)}</div>`+live(list).map(one).join("") : "").join("")
    : `<div class="grouphead">운동 구성</div>`+live(d.ex).map(one).join("");
  const holdBox = held.length ? `<details class="holdbox"><summary>보류 ${held.length}종목
    <i>· 세트 합계에서 빠짐</i></summary>${held.map(one).join("")}</details>` : "";
  const bar = (on && done!==d.total) ? rampBar(all,S,d.total,done,week) : "";
  /* 메인 리프트 카드의 본세트 수는 종목표 1번 행과 같은 값을 쓴다 */
  const mainRow = all.find(e=>e.n===1 && e.name.indexOf("본세트")>=0);
  /* 버튼에 적는 세트 수는 '오늘 할 세트'가 아니라 '타이머가 돌릴 세트'다.
     새 종목에 G·H열을 안 채우면 여기서 수가 어긋나 바로 눈에 띈다. */
  const tSets = timed.reduce((a,e)=>a+S(e),0);
  const go = timed.length
    ? `<button class="day-go" data-tm="">&#9654; 세션 타이머 시작 · ${tSets}세트${
        tSets!==done?` <i>(시간 없는 ${done-tSets}세트 제외)</i>`:""}</button>` : "";
  return `<h2 class="daytitle">${esc(d.t)}</h2><p class="daysub">${esc(d.sub)}</p>${bar}${go}
    ${d.main?liftCard(d.main,week,mainRow?S(mainRow):null):""}${notesHTML(d.pre)}${body}
    <div class="total"><span>세트 합계</span><b class="num">${done!==d.total?`${done} <i style="font-style:normal;font-weight:600;color:var(--mute);font-size:13px">/ 설계 ${d.total}</i>`:d.total}</b></div>${holdBox}${notesHTML(d.notes)}`;
}
/* ── 달력 ─────────────────────────────────────────────────────────
   8일 주기는 요일과 어긋나므로 '오늘 몇 일차인지'가 날짜로만 풀린다.
   기준점은 엑셀 증량기록의 '날짜 (1일차)'. 하루 빠뜨려 밀렸으면
   앱에서 '오늘 = N일차'로 다시 맞춘다 — 그 값은 이 폰에만 저장된다. */
const DOW="일월화수목금토";
const ymd = d => d.getFullYear()+"-"+String(d.getMonth()+1).padStart(2,"0")
                 +"-"+String(d.getDate()).padStart(2,"0");
const ymdParse = t => { const p=t.split("-").map(Number); return new Date(p[0],p[1]-1,p[2]); };
const dayGap = (a,b) => Math.round((ymdParse(a)-ymdParse(b))/86400000);
const AKEY="wk-anchor";
let ANCHOR=D.anchor||null;
try{ const o=JSON.parse(localStorage.getItem(AKEY)||"null"); if(o&&o.date) ANCHOR=o; }catch(e){}
const saveAnchor = () => { try{ localStorage.setItem(AKEY,JSON.stringify(ANCHOR)); }catch(e){} };

/* 날짜 → 그 날의 일차와 사이클 */
function cycleOf(t){
  if(!ANCHOR) return null;
  const L=D.homeOrder.length, n=dayGap(t,ANCHOR.date);
  return {day: D.homeOrder[((n%L)+L)%L],
          cycle: ANCHOR.cycle + Math.floor(n/L)};
}
const PLANMAP={}; D.plan.rows.forEach(([d,c,n])=>{ PLANMAP[d]={c:c,n:n}; });
const shortLabel = d => (PLANMAP[d]?PLANMAP[d].c:d).replace(/\s*\([^)]*\)/g,"").trim();
/* 일차별 색. 메인 리프트가 있는 날은 그 종목 색(빨강 스쿼트 · 파랑 데드 ·
   노랑 벤치 · 초록 밀리터리)을 쓰고, 없는 날은 남은 색을 순서대로 받는다.
   이웃한 일차끼리 색이 겹치지 않는 것이 목적이다 — 8일차 다음이 1일차라
   둘이 같은 색이면 사이클이 넘어간 건지 알 수 없다. */
const EXTRA=["#2BA6BF","#8A6CE0","#CE6BA5","#B08442"];
const DAYACC={};
(function(){ let i=0;
  D.homeOrder.forEach(d=>{
    const h=D.home[d];
    if(!h||h.rest) return;
    DAYACC[d] = (h.main&&LIFT[h.main]) ? LIFT[h.main].accent : EXTRA[i++%EXTRA.length];
  });
})();
const accentOf = d => DAYACC[d]||"";

let calYM=null;
function calHTML(){
  const now=new Date(), tKey=ymd(now), cur=cycleOf(tKey);
  const ym=calYM||{y:now.getFullYear(),m:now.getMonth()};
  const first=new Date(ym.y,ym.m,1), last=new Date(ym.y,ym.m+1,0);

  const head = cur
    ? `<div class="cal-now"><b>${cur.cycle}사이클 ${cur.day}일차 · ${esc(PLANMAP[cur.day]?PLANMAP[cur.day].c:"")}</b>
        <p>오늘 ${now.getMonth()+1}월 ${now.getDate()}일 (${DOW[now.getDay()]})${
          PLANMAP[cur.day]&&PLANMAP[cur.day].n?` · ${PLANMAP[cur.day].n}세트`:""}${
          D.home[cur.day]&&D.home[cur.day].main
            ? ` · ${esc(LIFT[D.home[cur.day].main].ko)} ${fmt(workWeight(D.home[cur.day].main,Math.min(WEEKS,Math.max(1,cur.cycle))))}kg`:""}</p>
       ${anchorFix(cur)}</div>`
    : `<div class="cal-now"><b>기준 날짜가 없다</b>
        <p>오늘이 몇 일차인지 한 번만 눌러주면 그 뒤로는 날짜로 자동 계산된다.
           (엑셀 증량기록 '날짜 (1일차)' 열에 적어두면 폰을 바꿔도 유지된다)</p>
       ${anchorFix(null)}</div>`;

  const cells=[];
  for(let i=0;i<first.getDay();i++) cells.push(`<div class="cal-cell pad"></div>`);
  for(let dnum=1;dnum<=last.getDate();dnum++){
    const d=new Date(ym.y,ym.m,dnum), key=ymd(d);
    /* 기준일 이전은 계산하지 않는다 — 8일 주기를 돌기 전이라 그때의 일차는 이 표가 알 수 없다 */
    const c=(ANCHOR&&dayGap(key,ANCHOR.date)>=0)?cycleOf(key):null;
    const dow=d.getDay(), lab=c?shortLabel(c.day):"";
    const rest=c&&PLANMAP[c.day]&&PLANMAP[c.day].n===0;
    const acc=c?accentOf(c.day):"";
    cells.push(`<div class="cal-cell${c?"":" blank"}${key===tKey?" today":""}${rest?" rest":""}${dow===0?" sun":""}"
      ${c?`data-cal="${c.day}" data-cyc="${c.cycle}"`:""} style="${acc?`--acc:${acc}`:""}">
      <b>${dnum}</b>${c?`<i${acc?` style="color:${acc}"`:""}>${c.day}일차</i>
      <span>${esc(lab)}</span>${c.day===D.homeOrder[0]?`<em>${c.cycle}사이클</em>`:""}`:""}</div>`);
  }

  return `<h2 class="daytitle">달력</h2>
  <p class="daysub">8일 주기는 요일과 어긋난다 — 오늘 무엇을 하는지는 날짜가 정한다. 칸을 누르면 그 일차로 간다.</p>
  ${head}
  <div class="cal-bar"><button id="cal-prev">‹</button>
    <h3>${ym.y}년 ${ym.m+1}월</h3>
    <button id="cal-today">오늘</button><button id="cal-next">›</button></div>
  <div class="cal-grid">${[...DOW].map((w,i)=>`<div class="cal-dow${i===0?" sun":""}">${w}</div>`).join("")}
    ${cells.join("")}</div>
  <div class="cal-key">${D.homeOrder.map(d=>{
    const h=D.home[d], acc=accentOf(d);
    return `<span><i style="background:${acc||"transparent"};${
      acc?"":"border:1px dashed var(--line)"}"></i>${d} ${esc(shortLabel(d))}${
      h&&h.main&&LIFT[h.main]?` <b>${esc(LIFT[h.main].ko)}</b>`:""}</span>`;
  }).join("")}</div>`;
}
function anchorFix(cur){
  return `<details class="cal-fix"><summary>${cur?"주기가 밀렸다면 — 오늘을 다른 일차로":"오늘은 몇 일차인가"}</summary>
    <div class="cal-days">${D.homeOrder.map(d=>
      `<button data-setday="${d}"${cur&&cur.day===d?' class="on"':""}>${d}일차 ${esc(shortLabel(d))}</button>`).join("")}</div>
    ${D.anchor?`<div class="cal-days"><button data-setday="__xl">엑셀 기준으로 되돌리기</button></div>`:""}</details>`;
}
/* 오늘을 N일차로 맞춘다 — 기준 날짜를 거꾸로 계산해 저장한다 */
function setToday(d){
  const L=D.homeOrder.length, i=D.homeOrder.indexOf(d);
  if(i<0) return;
  const back=new Date(); back.setDate(back.getDate()-i);
  const cur=cycleOf(ymd(new Date()));
  ANCHOR={date:ymd(back), cycle:(cur?cur.cycle:(D.config.week||1))};
  saveAnchor();
}

/* 검산 — 재택과 복귀가 다른 표를 쓴다. 복귀판은 주 단위라 환산이 없고,
   대신 현장 일이 채워주는 정도를 '보정' 열로 함께 보여준다. */
function checkHTML(){
  const back = mode==="back", rows = back ? D.backCheck : D.check;
  if(!rows||!rows.length) return `<h2 class="daytitle">세트 검산</h2><p class="daysub">표가 없습니다.</p>`;
  const adj = back && rows.some(r=>r[7]);
  return `<h2 class="daytitle">세트 검산</h2>
  <p class="daysub">${back
    ? "세트는 근육군 단위로 센다. 복귀판은 주 4회 · 주 단위라 이 숫자가 곧 주간 세트다. 실질 = 직접 + 간접 추정."
    : "세트는 근육군 단위로 센다. 실질 = 직접 + 간접 추정."}</p>
  <table class="tbl"><thead><tr><th>근육군</th><th>직접</th><th>간접</th><th>실질</th><th>판정</th>${adj?"<th>보정</th>":""}</tr></thead><tbody>
  ${rows.map(([m,dr,i,t,tag,v,,g])=>`<tr><td>${esc(m)}</td><td class="num">${dr}</td>
    <td class="num" style="color:var(--mute)">${i}</td><td class="num" style="font-weight:800">${t}</td>
    <td><span class="judge j-${tag}">${esc(v)}</span></td>${adj?`<td class="num" style="color:var(--mute)">${esc(g||"—")}</td>`:""}</tr>`).join("")}
  <tr><td style="font-weight:800">직접 세트 총계</td><td class="num" style="font-weight:800">${rows.reduce((a,r)=>a+r[1],0)}</td><td colspan="${adj?4:3}"></td></tr>
  </tbody></table>
  ${adj?`<div class="notes"><h5>현장 일 보정</h5><ul>
    <li><b style="color:var(--ink)">높음</b> — 현장 일이 하루치를 얹는다. 판정이 하한 미달이어도 실제로는 채워진다.</li>
    <li><b style="color:var(--ink)">중간</b> — 부분적으로만 채워진다. 판정을 절반쯤 믿는다.</li>
    <li><b style="color:var(--ink)">없음</b> — 현장 일이 안 건드린다. 판정을 액면대로 믿어야 하고, 볼륨을 쓸 곳도 여기다.</li></ul></div>`:""}
  <div class="notes"><h5>간접 출처</h5><ul>${rows.filter(r=>r[6]&&r[6]!=="—")
    .map(r=>`<li><b style="color:var(--ink)">${esc(r[0])}</b> — ${esc(r[6])}</li>`).join("")}</ul></div>
  ${!back&&RAMP?`<p class="footnote">이 표는 설계 볼륨 기준이다 — 램프가 끝난 ${RAMP.rows[RAMP.rows.length-1].week}주차부터의 값.
    이번 주에 실제로 하는 세트는 요일 탭과 주간 탭에서 본다.</p>`:""}`;
}
function rampHTML(){
  if(!RAMP) return "";
  const now=rampRow(week);
  return `<h3 class="sub-h">볼륨 램프 — 종목 복귀</h3><p class="daysub">${esc(RAMP.sub||"")}</p>
  <table class="tbl"><thead><tr><th>주차</th><th>총 세트</th><th>증가폭</th><th class="wrap">이 주차에 복귀하는 종목</th></tr></thead><tbody>
  ${RAMP.rows.map(r=>`<tr style="${r===now?'background:var(--raised)':''}">
    <td>${esc(r.label)}${r===now?' <span class="judge j-lo">현재</span>':''}</td>
    <td class="num" style="font-weight:800">${r.total}</td>
    <td class="num" style="color:var(--mute)">${esc(r.delta)}</td>
    <td class="wrap" style="color:var(--mute)">${esc(r.items)}</td></tr>`).join("")}
  </tbody></table>${notesHTML(RAMP.notes)}`;
}
function planHTML(){
  const p=D.plan;
  /* 요일 카드와 같은 방식으로 다시 센다 — 종목 단위로 빠지므로 합계에 계수를 곱하는 식이 아니다 */
  const wkSets = d => { const day=D.home[d];
    return (!day||day.rest) ? 0 : allEx(day).reduce((a,e)=>a+setsOf(e,week),0); };
  const sum=rampOn()?p.rows.reduce((a,[d,,n])=>a+(n?wkSets(d):0),0):p.total;
  const on=rampOn()&&sum!==p.total;
  const rows=p.rows.map(([d,c,n])=>`<tr class="${n?'':'rest'}"><td>${esc(d)}</td><td class="wrap">${esc(c)}</td>
    <td class="num">${n||"—"}</td>${on?`<td class="num" style="font-weight:800">${n?wkSets(d):"—"}</td>`:""}</tr>`).join("");
  return `<h2 class="daytitle">8일 주기 계획</h2><p class="daysub">${esc(p.sub)}</p>
  ${on?`<div class="rampbar"><b>볼륨 ${vWeek(week)}주차 · ${sum}세트 <span>/ 설계 ${p.total}</span></b>
      <span>오른쪽 열이 이번 주에 실제로 하는 세트다. 중량은 ${week}주차 그대로</span></div>`:""}
  <table class="tbl"><thead><tr><th>일차</th><th style="text-align:left">내용</th><th>${on?"설계":"세트"}</th>${on?"<th>이번 주</th>":""}</tr></thead><tbody>
  ${rows}
  <tr><td style="font-weight:800">합계</td><td></td><td class="num" style="font-weight:800">${p.total}</td>${on?`<td class="num" style="font-weight:800">${sum}</td>`:""}</tr>
  </tbody></table>${rampHTML()}${notesHTML(p.notes)}`;
}
function backPlanHTML(){
  const p=D.backPlan;
  return `<h2 class="daytitle">복귀 주간 계획</h2><p class="daysub">${esc(p.sub)}</p>
  <table class="tbl"><thead><tr><th>슬롯</th><th>근무</th><th>고정</th><th style="text-align:left">훈련</th><th>세트</th></tr></thead><tbody>
  ${p.rows.map(r=>`<tr class="${r[4]?'':'rest'}"><td>${esc(r[0])}</td>
    <td class="num" style="color:var(--mute)">${esc(r[1])}</td><td class="num" style="color:var(--mute)">${esc(r[2])}</td>
    <td class="wrap" style="font-weight:600;color:var(--ink)">${esc(r[3])}</td><td class="num">${r[4]||"—"}</td></tr>`).join("")}
  <tr><td style="font-weight:800">합계</td><td colspan="3"></td><td class="num" style="font-weight:800">${p.total}</td></tr>
  </tbody></table>${notesHTML(p.notes)}`;
}
function progressHTML(week){
  const keys=Object.keys(LIFT), rows=[];
  for(let w=1;w<=WEEKS;w++){
    rows.push(`<tr style="${w===week?'background:var(--raised)':''}"><td class="num">${w}주</td>`+
      keys.map(k=>`<td class="num">${fmt(workWeight(k,w))}</td>`).join("")+`</tr>`);
  }
  return `<h2 class="daytitle">증량 계획</h2>
  <p class="daysub">목표 반복을 못 채운 주가 없다고 가정한 계획선. 실제 진행은 증량기록 시트가 결정한다.</p>
  <table class="tbl"><thead><tr><th>주차</th>${keys.map(k=>`<th>${esc(LIFT[k].ko)}</th>`).join("")}</tr></thead>
  <tbody>${rows.join("")}</tbody></table>
  <div class="notes"><h5>주당 증량폭</h5><ul>${keys.map(k=>
    `<li>${esc(LIFT[k].ko)} ${fmt(LIFT[k].inc)}kg</li>`).join("")}</ul></div>
  ${notesHTML(D.progressNotes)}`;
}

/* ── 중량 편집 UI ── 숫자를 탭 → 입력 → 확인(엔터·바깥 탭)으로 저장.
   빈 값으로 확인하면 엑셀 값으로 되돌아간다. */
function grabInput(el, value, numeric, onCommit){
  if(el.querySelector("input")) return;
  el.innerHTML=`<input class="w-input" type="text"${numeric?' inputmode="decimal"':''}`+
    ` value="${escAttr(value)}" placeholder="—">`;
  const inp=el.querySelector("input");
  inp.focus(); inp.select();
  let done=false;
  const finish=commit=>{ if(done) return; done=true; if(commit) onCommit(inp.value.trim()); render(); };
  inp.onblur=()=>finish(true);
  inp.onkeydown=ev=>{ if(ev.key==="Enter") finish(true); else if(ev.key==="Escape") finish(false); };
}
function editLift(el){
  const k=el.dataset.lift;
  grabInput(el, fmt(workWeight(k,week)), true, v=>{
    const n=parseFloat(v);
    if(v===""){ delete OV.lifts[k]; }
    else if(!isNaN(n)&&n>0){
      // 이번 주 무게를 입력받아 시작중량을 역산 — 이후 주차·원판·웜업이 전부 따라온다
      const start=n-LIFT[k].inc*(week-1);
      if(Math.abs(start-LIFT[k].start)<0.001) delete OV.lifts[k];
      else OV.lifts[k]={start:start, base:LIFT[k].start};
    }
    saveOV();
  });
}
function editEx(el){
  const key=el.dataset.exkey, e=exByKey(key);
  if(!e) return;
  const cur=OV.ex[key]?OV.ex[key].w:e.w;
  grabInput(el, cur==="—"?"":cur, false, v=>{
    if(!v||v===e.w) delete OV.ex[key];
    else OV.ex[key]={w:v, base:e.w};
    saveOV();
  });
}
function ovHTML(){
  const items=[];
  for(const k in OV.lifts)
    items.push(`${LIFT[k].ko} — 시작중량 ${fmt(OV.lifts[k].base)} → ${fmt(OV.lifts[k].start)}kg`);
  for(const k in OV.ex){
    const p=k.split("|");
    items.push(`${p[0]==="back"?"복귀 ":""}${p[1]} · ${p[2]} — ${OV.ex[k].base} → ${OV.ex[k].w}`);
  }
  if(!items.length) return "";
  return `<div class="ovbox"><h5>이 폰에서 수정한 값 <span>· 엑셀 원본은 그대로이니 나중에 옮겨 적으세요</span></h5>
    <ul>${items.map(t=>`<li>${esc(t)}</li>`).join("")}</ul>
    <button id="ov-reset">전부 엑셀 값으로 되돌리기</button></div>`;
}
/* ── 인터벌 타이머 ────────────────────────────────────────────────
   엑셀 G·H열(수행 초 · 휴식 초)을 그대로 읽어 종목 → 세트 순으로 걷는다.
   남은 시간은 setInterval 누적이 아니라 목표 시각과의 차로 낸다 —
   화면이 꺼져 콜백이 밀려도 돌아왔을 때 숫자가 어긋나지 않는다.
   화면 잠금은 Wake Lock으로 막는다. 브라우저가 백그라운드에서 소리를
   막을 수 있으므로 화면을 켜둔 채 쓰는 것이 전제다. */
const PH={prep:["준비","var(--bp)"],work:["운동","var(--mp)"],rest:["휴식","var(--sq)"]};
const PREP=5;
const mmss=v=>{const m=Math.floor(v/60),x=v%60;return (m<10?"0":"")+m+":"+(x<10?"0":"")+x;};
let TM=null, tmInt=null, tmAC=null, tmWL=null, tmLeft=-1;

function tmSteps(d,week,from){
  const on=rampOn(), S=e=>on?setsOf(e,week):e.s;
  const live=allEx(d).filter(e=>e.s!==0&&S(e)>0&&e.wk>0);
  const st=[];
  live.forEach(e=>{ const n=S(e);
    for(let k=1;k<=n;k++){
      st.push({p:"work",t:e.wk,e:e,set:k,sets:n});
      if(e.rt>0) st.push({p:"rest",t:e.rt,e:e,set:k,sets:n});
    }});
  while(st.length&&st[st.length-1].p==="rest") st.pop();   /* 세션 끝의 휴식은 뺀다 */
  let cut=0;
  if(from){ const j=st.findIndex(x=>x.e.name===from); if(j>0) cut=j; }
  const out=st.slice(cut);
  if(out.length) out.unshift({p:"prep",t:PREP,e:out[0].e,set:out[0].set,sets:out[0].sets});
  out.forEach(x=>{ x.li=live.indexOf(x.e)+1; x.ln=live.length; });
  return out;
}
function tmW(e){
  const d=(mode==="home"?D.home:D.back)[tab];
  if(e.w==="위 참조"&&d&&d.main) return fmt(workWeight(d.main,week))+"kg";
  const ov=OV.ex[mode+"|"+tab+"|"+e.name], w=String(ov?ov.w:e.w);
  return /^[\d.~\s]+$/.test(w) ? w.replace(/\s+$/,"")+"kg" : w;
}
/* -- 화면이 꺼져도 알림이 울리게 하는 두 가지 --------------------
   (1) 무음 유지음: 사람 귀에 안 들리는 아주 작은 소리를 계속 흘린다.
       크롬은 '소리를 내는 탭'을 얼리지 않고, 타이머도 분당 1회까지 늦추지 않는다.
       화면을 꺼도 오디오 스레드는 계속 돈다 — 웹 라디오가 꺼진 화면에서
       계속 나오는 것과 같은 원리다.
   (2) 미리 예약: 알림음을 JS가 그 시각에 깨어나서 울리는 게 아니라
       오디오 시계에 90초 앞까지 미리 걸어둔다. JS가 얼마나 늦어지든
       예약된 소리는 하드웨어가 정시에 낸다.
   진동은 예약이 안 되고 화면이 꺼져 있으면 브라우저가 막으므로 보조 수단이다. */
const LOOKAHEAD=90000;
let tmKeep=null, tmPend=[], tmSchedAt=0, tmIdle=null;

/* 말소리도 삐 소리와 같은 오디오 시계에 건다.
   기기 TTS(speechSynthesis)를 안 쓰는 이유: 화면이 꺼지면 크롬이
   not-allowed로 통째로 거부한다 — 정작 필요한 순간에 안 나온다.
   미리 구운 wav라 화면이 꺼져 있어도 예약대로 난다. */
/* 말은 한 파일에 죽 이어 붙여 두고(스프라이트) 필요한 구간만 잘라 쓴다.
   조각을 파일로 두면 종목 이름·숫자까지 200개가 넘어 첫 방문에 요청이
   200번 난다. V.map["랫풀다운"] = [시작초, 길이초] — 그게 전부다. */
const V=D.voice||null;
let vBuf=null, vBusy=false;
let voiceOn=true; try{ voiceOn=localStorage.getItem("wk-voice")!=="0"; }catch(e){}
function tmVoiceLoad(){
  if(!tmAC||!V||vBuf||vBusy) return;
  vBusy=true;
  fetch(V.src).then(r=>r.arrayBuffer()).then(b=>tmAC.decodeAudioData(b))
    .then(buf=>{ vBuf=buf; if(TM&&TM.run) tmSchedule(); })
    .catch(()=>{ vBusy=false; });                    /* 못 받으면 삐 소리만 */
}
const tmHas = k => !!(V&&V.map[k]);
const tmDur = k => tmHas(k) ? V.map[k][1] : 0;
function tmSay(kind,at){
  if(!voiceOn||!vBuf||!tmHas(kind)) return null;
  const sl=V.map[kind];
  try{
    const src=tmAC.createBufferSource();
    src.buffer=vBuf; src.connect(tmAC.destination);
    src.start(at,sl[0],sl[1]);
    return src;
  }catch(e){ return null; }
}
/* 조각을 차례로 건다. ","는 쉼표 — 소리는 없고 사이만 벌린다.
   낱말 사이를 너무 붙이면 한 덩어리로 뭉쳐 들려 무슨 말인지 안 잡힌다. */
const VGAP=.13, VPAUSE=.3;
function tmSayAll(keys,at){
  const n=[]; let t=at;
  keys.forEach(k=>{
    if(k===","){ t+=VPAUSE; return; }
    const s=tmSay(k,t);
    if(s){ n.push(s); t+=tmDur(k)+VGAP; }
  });
  return n;
}

/* -- 숫자를 조각 이름으로 --------------------------------------
   중량은 주차마다 오르고 앱에서 고칠 수도 있어서 문장을 미리 구워 둘 수가
   없다. 그래서 0~99와 백 단위만 구워 두고 여기서 이어 붙인다.
   규칙은 voice_build.py의 sino()/num_keys()와 한 글자도 다르면 안 된다 —
   한쪽만 고치면 앱이 조각을 못 찾아 그 부분이 통째로 조용해진다. */
const KO1=["","일","이","삼","사","오","육","칠","팔","구"];
const KOF={"5":"점오","25":"점이오","75":"점칠오","125":"점일이오"};
function koTens(n){
  if(n===0) return "영";
  const t=Math.floor(n/10), o=n%10;
  return (t===0?"":t===1?"십":KO1[t]+"십")+KO1[o];
}
function koNum(s){
  const m=/^(\d+)(?:\.(\d+))?$/.exec(String(s).trim());
  if(!m) return null;
  const n=+m[1], f=(m[2]||"").replace(/0+$/,"");
  if(n>=1000) return null;
  const k=[], h=Math.floor(n/100);
  if(h){ k.push((h===1?"":KO1[h])+"백"); if(n%100) k.push(koTens(n%100)); }
  else k.push(koTens(n));
  if(f){
    if(KOF[f]) k.push(KOF[f]);
    else { k.push("점"); for(const c of f) k.push(koTens(+c)); }
  }
  return k;
}
/* '40' → 사십 킬로 · '30~35' → 삼십 에서 삼십오 킬로 · 숫자가 아니면 null */
function koRange(s,unit){
  const p=String(s).trim().split("~");
  if(p.length>2) return null;
  const a=koNum(p[0]); if(!a) return null;
  if(p.length===1) return a.concat([unit]);
  const b=koNum(p[1]); if(!b) return null;
  return a.concat(["에서"],b,[unit]);
}
/* 숫자 꼴이면 조립하고, 아니면 통째로 구워 둔 칸('맨몸')을 찾는다 */
function koCell(s,unit){
  return koRange(s,unit) || (tmHas(s)?[s]:[]);
}
/* 다음 종목 안내. 값은 그 순간 화면에 떠 있는 것을 그대로 읽으므로
   주차 증량도 앱에서 고친 중량도 따로 손댈 데가 없다. */
function tmNextKeys(e,sets){
  const k=["다음 종목",",",e.name];
  if(e.mg) k.push(",",e.mg);
  k.push(",");
  k.push.apply(k,(koNum(String(sets))||[]).concat(["세트"]));
  k.push(",");
  k.push.apply(k,koCell(String(e.r||""),"회"));
  k.push(",");
  k.push.apply(k,koCell(tmW(e).replace(/kg$/,""),"킬로"));
  return k.filter(x=>x===","||tmHas(x));
}
const tmLen = keys => keys.reduce(
  (a,k)=>a+(k===","?VPAUSE:tmDur(k)+VGAP), 0);
/* 이 휴식 다음에 종목이 바뀔 때만 읊는다. 같은 종목의 다음 세트로
   이어질 뿐이면 조용하다 — 매 세트 떠들면 그냥 소음이다.
   휴식이 끝나갈 때 나는 3·2·1 딸깍 소리에 말이 묻히지 않도록,
   그 전에 다 끝낼 수 있을 때만 입을 뗀다. 5초짜리 '준비'에서 읊지 않는
   것도 같은 이유다 — 시작 버튼을 누른 그 순간엔 화면을 보고 있다. */
function tmNextSay(idx,at,off){
  if(!TM||idx==null||!voiceOn||!vBuf) return [];
  const cur=TM.st[idx], nx=TM.st[idx+1];
  if(!cur||!nx||nx.e===cur.e) return [];
  const k=tmNextKeys(nx.e,nx.sets);
  if(off+tmLen(k)>cur.t-4) return [];      /* 딸깍 소리까지 4초는 비워 둔다 */
  return tmSayAll(k,at);
}

const tmVib=p=>{ try{ if(document.visibilityState==="visible"&&navigator.vibrate) navigator.vibrate(p); }catch(e){} };

function tmKeepStart(){
  try{
    const b=tmAC.createBuffer(1,tmAC.sampleRate,tmAC.sampleRate), ch=b.getChannelData(0);
    for(let i=0;i<ch.length;i++) ch[i]=(Math.random()*2-1)*2e-4;   /* -74dB — 안 들린다 */
    tmKeep=tmAC.createBufferSource(); tmKeep.buffer=b; tmKeep.loop=true;
    tmKeep.connect(tmAC.destination); tmKeep.start();
  }catch(e){}
}
/* quiet=true — 컨텍스트만 만들고 말소리를 미리 디코딩한다. 소리는 안 낸다.
   페이지가 열릴 때 한 번 이렇게 불러두면 첫 '준비'부터 말이 나온다. */
function tmAudio(quiet){
  clearTimeout(tmIdle);
  if(!tmAC){
    try{ tmAC=new (window.AudioContext||window.webkitAudioContext)(); }
    catch(e){ tmAC=null; return null; }
    tmVoiceLoad();
  }
  if(!quiet){
    if(tmAC.state==="suspended") tmAC.resume();
    if(!tmKeep) tmKeepStart();
    tmVoiceLoad();
  }
  return tmAC;
}
function tmTone(f,dur,vol,at){
  try{
    const o=tmAC.createOscillator(), g=tmAC.createGain();
    o.type="sine"; o.frequency.value=f;
    g.gain.setValueAtTime(vol,at); g.gain.exponentialRampToValueAtTime(.0001,at+dur);
    o.connect(g); g.connect(tmAC.destination);
    o.start(at); o.stop(at+dur+.05);
    return o;
  }catch(e){ return null; }
}
/* 삐 소리와 말 사이의 간격. 삐 소리는 여운이 길어서 바짝 붙이면 말이
   그 위에 얹혀 안 들린다 — 소리가 잦아든 뒤에 입을 뗀다. */
const VAFTER=.36;
function tmCueAt(kind,wall,idx){
  if(!tmAC) return;
  const at=tmAC.currentTime+Math.max(0,(wall-Date.now())/1000), n=[];
  let say=[];
  /* 삐 소리가 먼저 귀를 잡고, 뒤이어 무엇인지 말한다 */
  if(kind==="work"){ n.push(tmTone(1046,.35,.6,at)); n.push(tmSay("work",at+.35+VAFTER)); }
  else if(kind==="rest"){ const beep=.22+.15;   /* 두 번째 삐 소리가 끝나는 시각 */
                          n.push(tmTone(660,.15,.5,at)); n.push(tmTone(660,.15,.5,at+.22));
                          n.push(tmSay("rest",at+beep+VAFTER));
                          /* 다음이 새 종목이면 휴식 시작에 한 번 읊는다 —
                             쉬는 동안 옮겨 갈 기구와 무게를 미리 알라고 */
                          const off=beep+VAFTER+tmDur("rest")+.5;
                          say=tmNextSay(idx,at+off,off); }
  else if(kind==="prep"){ n.push(tmTone(784,.12,.4,at)); n.push(tmSay("prep",at+.12+VAFTER)); }
  else if(kind==="tick") n.push(tmTone(880,.07,.3,at));
  else { n.push(tmTone(1046,.5,.6,at)); n.push(tmSay("done",at+.5+VAFTER)); }
  tmPend.push({at:wall,kind:kind,nodes:n.filter(Boolean).concat(say),say:say});
}
/* 읊는 중인 안내를 끊는다. 건너뛰기로 다른 종목에 가 놓고 이전 안내를
   계속 듣고 있으면 틀린 무게를 듣게 된다. 삐 소리는 짧아서 그냥 둔다. */
function tmHush(){
  tmPend.forEach(x=>(x.say||[]).forEach(o=>{ try{ o.stop(0); }catch(e){} }));
}
/* 아직 안 난 예약을 모두 취소한다. 나고 있는 소리는 자르지 않는다 */
function tmDrop(){
  const now=Date.now(), keep=[];
  tmPend.forEach(x=>{
    if(x.at<=now+250){ if(x.at>now-8000) keep.push(x); return; }
    x.nodes.forEach(o=>{ try{ o.stop(0); }catch(e){} });
  });
  tmPend=keep;
}
/* 지금부터 90초 안에 울릴 것을 전부 오디오 시계에 건다 */
function tmSchedule(){
  tmDrop(); tmSchedAt=Date.now();
  if(!TM||!TM.run||!tmAC) return;
  const now=Date.now(), lim=now+LOOKAHEAD;
  const add=(k,at,idx)=>{ if(at>now+250&&at<=lim) tmCueAt(k,at,idx); };
  let end=TM.endAt;
  for(let i=TM.i;i<TM.st.length;i++){
    if(i>TM.i) end+=TM.st[i].t*1000;
    for(let k=3;k>=1;k--) add("tick",end-k*1000);
    /* end에 시작되는 건 i+1번 구간이다 — 안내는 그 구간을 기준으로 만든다 */
    add(i+1<TM.st.length?TM.st[i+1].p:"done",end,i+1);
    if(end>lim) break;
  }
}
/* 방금 이 구간에 들어왔다 — 예약이 아니라 즉시 */
function tmEnter(){
  if(!TM) return;
  const p=TM.st[TM.i].p;
  tmHush();
  tmCueAt(p,Date.now()+20,TM.i);
  tmVib(p==="work"?[350]:p==="rest"?[140,90,140]:[90]);
  tmSchedule();
}
/* 잠금화면·알림창에 종목과 세트를 띄우고 거기서 일시정지할 수 있게 한다 */
function tmMedia(){
  if(!("mediaSession" in navigator)||!TM) return;
  const s=TM.st[TM.i];
  try{
    navigator.mediaSession.metadata=new MediaMetadata({
      title:s.e.name, artist:PH[s.p][0]+" · 세트 "+s.set+" / "+s.sets, album:TM.title});
    navigator.mediaSession.playbackState=TM.run?"playing":"paused";
    navigator.mediaSession.setActionHandler("play",tmPause);
    navigator.mediaSession.setActionHandler("pause",tmPause);
    navigator.mediaSession.setActionHandler("nexttrack",()=>tmGo(1));
    navigator.mediaSession.setActionHandler("previoustrack",()=>tmGo(-1));
  }catch(e){}
}
async function tmLock(){ try{ if("wakeLock" in navigator && !tmWL) tmWL=await navigator.wakeLock.request("screen"); }catch(e){} }
async function tmUnlock(){ try{ if(tmWL){ const w=tmWL; tmWL=null; await w.release(); } }catch(e){} }
document.addEventListener("visibilitychange",()=>{
  if(document.visibilityState!=="visible"||!TM) return;
  /* 돌아왔다 — 화면을 지금 시각에 맞추고 예약도 다시 건다 */
  if(tmAC&&tmAC.state==="suspended") tmAC.resume();
  if(TM.run){ tmLock(); tmLoop(); tmSchedule(); }
});

/* 종목 한 줄 요약. 어느 근육을 쓰는지 · 몇 kg · 몇 회 — 상단과 하단이
   같은 함수를 쓰므로 한쪽만 형식이 어긋날 일이 없다. */
const tmMeta = e => [e.mg, tmW(e), e.r+"회"].filter(Boolean).join(" · ");
function tmPaint(){
  if(!TM) return;
  const s=TM.st[TM.i], nx=TM.st[TM.i+1], ph=PH[s.p];
  document.getElementById("tm-dial").style.setProperty("--phase",ph[1]);
  document.getElementById("tm-phase").textContent=ph[0];
  document.getElementById("tm-ctx").textContent=TM.title+" · "+s.li+"/"+s.ln+" 종목";
  document.getElementById("tm-name").textContent=s.e.name;
  document.getElementById("tm-meta").textContent=tmMeta(s.e);
  document.getElementById("tm-set").textContent="세트 "+s.set+" / "+s.sets;
  document.getElementById("tm-time").textContent=
    mmss(TM.run?Math.max(0,Math.ceil((TM.endAt-Date.now())/1000)):Math.ceil(TM.rest/1000));
  document.getElementById("tm-play").textContent=TM.run?"일시정지":"계속";
  const up=document.getElementById("tm-up");
  if(nx&&nx.e!==s.e)
    /* 종목이 바뀐다 — 쉬는 동안 눈으로도 확인하게 상단과 같은 형식으로 적는다.
       읊어 주는 내용과 같다. 소리를 껐거나 못 들었을 때의 몫이기도 하다. */
    up.innerHTML="다음 ▸ <b>"+esc(nx.e.name)+"</b><i>"
                +esc(tmMeta(nx.e))+" · "+nx.sets+"세트</i>";
  else up.textContent = !nx ? "마지막 세트"
    : s.p==="prep" ? "이어서 ▸ 운동 "+mmss(nx.t)+" × "+nx.sets+"세트"
    : nx.p==="rest" ? "다음 ▸ 휴식 "+mmss(nx.t) : "다음 ▸ "+nx.set+"세트째";
}
function tmLoop(){
  if(!TM||!TM.run) return;
  const now=Date.now(); let moved=false;
  while(TM.endAt<=now&&TM.i<TM.st.length-1){ TM.i++; TM.endAt+=TM.st[TM.i].t*1000; moved=true; }
  if(TM.endAt<=now){ tmVib([300,120,300,120,300]); tmClose(); return; }
  if(moved){
    tmPaint(); tmMedia(); tmLeft=-1;
    tmVib(TM.st[TM.i].p==="work"?[350]:TM.st[TM.i].p==="rest"?[140,90,140]:[90]);
  }
  /* 예약은 5초마다 다시 건다. 구간 전환 때만 걸면 휴식 150초처럼
     예약 창(90초)보다 긴 구간의 알림이 통째로 빠진다. 다시 걸면서
     오디오 시계와 벽시계 사이의 드리프트도 같이 잡힌다. */
  if(moved||now-tmSchedAt>5000) tmSchedule();
  const left=Math.max(0,Math.ceil((TM.endAt-now)/1000));
  if(left!==tmLeft){
    tmLeft=left;
    document.getElementById("tm-time").textContent=mmss(left);
  }
}
function tmStart(d,from){
  const st=tmSteps(d,week,from);
  if(!st.length) return;
  tmAudio();
  TM={st:st,i:0,endAt:Date.now()+st[0].t*1000,rest:0,run:true,title:d.t};
  document.getElementById("tm").hidden=false;
  document.body.style.overflow="hidden";
  tmLock(); tmPaint(); tmMedia(); tmEnter(); tmLeft=-1;
  clearInterval(tmInt); tmInt=setInterval(tmLoop,120);
}
function tmGo(delta){
  if(!TM) return;
  const j=TM.i+delta;
  if(j<0||j>=TM.st.length) return;
  TM.i=j; TM.endAt=Date.now()+TM.st[j].t*1000; TM.rest=TM.st[j].t*1000;
  tmLeft=-1; tmPaint(); tmMedia();
  if(TM.run) tmEnter(); else tmDrop();
}
function tmPause(){
  if(!TM) return;
  if(TM.run){ TM.rest=Math.max(0,TM.endAt-Date.now()); TM.run=false; tmUnlock();
              tmHush(); tmDrop(); }
  else { TM.endAt=Date.now()+TM.rest; TM.run=true; tmLock(); tmSchedule(); }
  tmLeft=-1; tmPaint(); tmMedia();
}
function tmClose(){
  clearInterval(tmInt); tmInt=null; TM=null; tmLeft=-1; tmUnlock(); tmHush(); tmDrop();
  /* 유지음은 끊는다 — 타이머를 닫고도 계속 흘리면 배터리만 먹는다.
     다만 '완료' 말소리가 남아 있을 수 있어 2.5초 뒤에 끈다. */
  clearTimeout(tmIdle);
  tmIdle=setTimeout(()=>{ try{
    if(tmKeep){ tmKeep.stop(); tmKeep=null; }
    if(tmAC) tmAC.suspend();
  }catch(e){} }, 2500);
  try{ if("mediaSession" in navigator){ navigator.mediaSession.metadata=null;
    navigator.mediaSession.playbackState="none"; } }catch(e){}
  document.getElementById("tm").hidden=true;
  document.body.style.overflow="";
}

function bindEdits(){
  document.querySelectorAll("[data-lift]").forEach(el=>{ el.onclick=()=>editLift(el); });
  document.querySelectorAll("[data-exkey]").forEach(el=>{ el.onclick=()=>editEx(el); });
  const rb=document.getElementById("ov-reset");
  if(rb) rb.onclick=()=>{ OV={lifts:{},ex:{}}; saveOV(); render(); };
  const cp=document.getElementById("cal-prev"), cn=document.getElementById("cal-next"),
        ct=document.getElementById("cal-today");
  const shift=n=>{ const now=new Date(), b=calYM||{y:now.getFullYear(),m:now.getMonth()};
    const d=new Date(b.y,b.m+n,1); calYM={y:d.getFullYear(),m:d.getMonth()}; render(); };
  if(cp) cp.onclick=()=>shift(-1);
  if(cn) cn.onclick=()=>shift(1);
  if(ct) ct.onclick=()=>{ calYM=null; render(); };
  document.querySelectorAll("[data-setday]").forEach(el=>{ el.onclick=()=>{
    const v=el.dataset.setday;
    if(v==="__xl"){ ANCHOR=D.anchor; try{ localStorage.removeItem(AKEY); }catch(e){} }
    else setToday(v);
    render(); }; });
  document.querySelectorAll("[data-cal]").forEach(el=>{ el.onclick=()=>{
    const c=Number(el.dataset.cyc);
    if(c>=1&&c<=WEEKS) week=c;
    tab=el.dataset.cal; render(); window.scrollTo({top:0}); }; });
  /* data-tm="" = 세션 처음부터 · data-tm="종목명" = 그 종목부터 */
  document.querySelectorAll("[data-tm]").forEach(el=>{ el.onclick=()=>{
    const d=(mode==="home"?D.home:D.back)[tab];
    if(d&&!d.rest) tmStart(d,el.dataset.tm||null); }; });
}

let mode="home", week=D.config.week||1, tab=D.homeOrder[0];
const tabsFor = m => (m==="home"
  ? D.homeOrder.map(d=>[d,d+"일"]).concat([["cal","달력"],["plan","주기"],["prog","증량"],["check","검산"]])
  : D.backOrder.map(d=>[d,d]).concat([["bplan","주간"],["check","검산"]]));

function render(){
  const nav=document.getElementById("nav");
  nav.innerHTML=tabsFor(mode).map(([k,l])=>`<button data-k="${k}" aria-pressed="${k===tab}">${l}</button>`).join("");
  nav.querySelectorAll("button").forEach(b=>b.onclick=()=>{tab=b.dataset.k;render();window.scrollTo({top:0});});
  document.getElementById("weekbar").style.display = mode==="home"?"flex":"none";
  document.getElementById("wk-track").style.display = mode==="home"?"flex":"none";
  document.getElementById("wk-num").textContent=week;
  /* 큰 숫자는 중량 주차다. 볼륨이 그보다 앞서 있으면 여기서 같이 밝힌다 */
  document.getElementById("wk-of").textContent=`WEEK / ${WEEKS}`
    + (rampOn()&&vWeek(week)!==week ? ` · 볼륨 ${vWeek(week)}주` : "");
  document.getElementById("wk-track").innerHTML=
    Array.from({length:WEEKS},(_,i)=>`<span class="${i<week?'on':''}"></span>`).join("");
  document.getElementById("wk-down").disabled=week<=1;
  document.getElementById("wk-up").disabled=week>=WEEKS;
  const set = mode==="home"?D.home:D.back;
  let html;
  if(tab==="cal") html=calHTML();
  else if(tab==="check") html=checkHTML();
  else if(tab==="plan") html=planHTML();
  else if(tab==="prog") html=progressHTML(week);
  else if(tab==="bplan") html=backPlanHTML();
  else html = set[tab] ? dayHTML(set[tab],week) : planHTML();
  document.getElementById("app").innerHTML = html + ovHTML() +
    `<p class="footnote">원본: ${esc(D.source)} · 중량은 시트와 동일한 계산식(MROUND 2.5kg, 웜업 40/60/75/90%)으로 산출됩니다. 중량 숫자를 탭하면 이 폰에서만 고칠 수 있습니다.</p>`;
  bindEdits();
  try{ const c=cycleOf(ymd(new Date()));
    localStorage.setItem("wk-prog", JSON.stringify(
      {mode,week,tab,on:ymd(new Date()),cd:c?c.day:null})); }catch(e){}
}
const tmMuteBtn=document.getElementById("tm-mute");
const tmMutePaint=()=>{ tmMuteBtn.textContent=voiceOn?"\uD83D\uDD0A":"\uD83D\uDD07";
  tmMuteBtn.classList.toggle("off",!voiceOn); };
tmMutePaint();
try{ tmAudio(true); }catch(e){}     /* 말소리 미리 디코딩 — 출력은 없다 */
tmMuteBtn.onclick=()=>{
  voiceOn=!voiceOn;
  try{ localStorage.setItem("wk-voice",voiceOn?"1":"0"); }catch(e){}
  tmMutePaint();
  if(!voiceOn) tmHush();            /* 끄면 하던 말도 그 자리에서 멈춘다 */
  if(TM&&TM.run) tmSchedule();      /* 이미 걸어둔 예약을 새 설정으로 다시 건다 */
};
document.getElementById("tm-x").onclick=tmClose;
document.getElementById("tm-play").onclick=tmPause;
document.getElementById("tm-next").onclick=()=>tmGo(1);
document.getElementById("tm-prev").onclick=()=>tmGo(-1);
document.getElementById("ver").textContent=(D.source.match(/v\d+/)||[""])[0];
document.getElementById("wk-up").onclick=()=>{if(week<WEEKS){week++;render();}};
document.getElementById("wk-down").onclick=()=>{if(week>1){week--;render();}};
document.getElementById("m-home").onclick=()=>{mode="home";tab=D.homeOrder[0];
  document.getElementById("m-home").setAttribute("aria-pressed","true");
  document.getElementById("m-back").setAttribute("aria-pressed","false");render();};
document.getElementById("m-back").onclick=()=>{mode="back";tab=D.backOrder[0];
  document.getElementById("m-back").setAttribute("aria-pressed","true");
  document.getElementById("m-home").setAttribute("aria-pressed","false");render();};
try{
  const st=JSON.parse(localStorage.getItem("wk-prog")||"null");
  /* 오늘 할 일차로 간다 — 8일 주기는 요일로 못 세니까.
     날짜가 바뀌었을 때뿐 아니라 기준일이 수정돼 오늘 일차 자체가 달라졌을 때도 옮긴다
     (st.cd = 마지막으로 저장할 때의 오늘 일차. 없으면 옛 버전이므로 역시 옮긴다). */
  const cur0=cycleOf(ymd(new Date()));
  const moved = !st || st.on!==ymd(new Date()) || st.cd!==cur0.day;
  if(cur0&&moved&&(!st||st.mode!=="back")){
    tab=cur0.day; week=Math.min(WEEKS,Math.max(1,cur0.cycle));
  } else if(st&&st.mode){ mode=st.mode; week=Math.min(st.week||1,WEEKS); tab=st.tab||tab;
    document.getElementById("m-home").setAttribute("aria-pressed",mode==="home");
    document.getElementById("m-back").setAttribute("aria-pressed",mode==="back"); }
}catch(e){}
render();
</script>
</body>
</html>
"""


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="운동프로그램 워크북 → 모바일 HTML")
    ap.add_argument("xlsx", nargs="?", help="워크북 경로 (생략하면 폴더에서 자동 탐색)")
    ap.add_argument("-o", "--out", help="출력 HTML 경로 (기본: 같은 이름 _모바일.html)")
    a = ap.parse_args()
    if a.xlsx:
        src = Path(a.xlsx)
        if not src.exists():
            sys.exit(f"파일을 찾을 수 없습니다: {src}")
    else:
        src = pick_latest(Path(__file__).resolve().parent)
        print(f"워크북: {src.name}")
    out = Path(a.out) if a.out else src.with_name(src.stem + "_모바일.html")
    build(src, out)
